"""RFQGeneratorAgent
===================
Agent that generates RFQ (Request for Quotation) documents (Excel + PDF)
for a HVAC procurement request.

When the user selects "Use Recommendation" in the RFQ modal form, this agent
is invoked automatically with selection_mode="RECOMMENDED", which fetches the
latest RecommendationResult for the request and uses its system_type_code to
drive the document content.

Entry points
------------
    from apps.procurement.agents.RFQ_Generator_Agent import RFQGeneratorAgent

    # -- Triggered from form when user clicks "Use Recommendation" ----------
    result = RFQGeneratorAgent.run(
        proc_request,
        selection_mode="RECOMMENDED",   # "Use Recommendation" button in modal
        generated_by=request.user,
        save_record=True,               # persist GeneratedRFQ + upload blob
    )

    # -- Manual system selection (user picks a specific system code) --------
    result = RFQGeneratorAgent.run(
        proc_request,
        selection_mode="VRF",           # any HVACSystemType code
        qty_overrides={0: 3, 1: 5},     # optional row-index -> qty overrides
        generated_by=request.user,
        save_record=True,
    )

Returns
-------
    RFQResult dataclass with:
        xlsx_bytes      bytes             -- ready-to-serve Excel workbook
        pdf_bytes       bytes             -- ready-to-serve PDF (empty if reportlab missing)
        rfq_ref         str               -- e.g. "RFQ-0001-20260101"
        filename_xlsx   str
        filename_pdf    str
        system_code     str               -- e.g. "VRF"
        system_label    str               -- e.g. "VRF (Variable Refrigerant Flow)"
        selection_basis str               -- "AI / Rules Engine Recommendation" or "Manual Selection"
        confidence_pct  int               -- 0-100 (0 for manual)
        scope_rows      list of tuples    -- (category, description, unit, qty)
        rfq_record      GeneratedRFQ|None -- populated when save_record=True
        error           str|None          -- non-fatal error message if partial failure
"""
from __future__ import annotations

import datetime
import io
import json as _json
import logging
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Type aliases
# ---------------------------------------------------------------------------
ScopeRow = Tuple[str, str, str, Any]   # (category, description, unit, qty)

# ---------------------------------------------------------------------------
# System label registry
# ---------------------------------------------------------------------------
_SYSTEM_LABELS: Dict[str, str] = {
    "VRF":          "VRF (Variable Refrigerant Flow)",
    "SPLIT_AC":     "Split Air Conditioning System",
    "PACKAGED_DX":  "Packaged DX Unit",
    "FCU":          "Fan Coil Unit (FCU)",
    "CHILLER":      "Chilled Water System (Chiller)",
    "CASSETTE":     "Cassette Type Split AC",
    "AHU":          "Air Handling Unit (AHU)",
    "DUCTING":      "Ducting & Accessories",
}

# Keyword -> canonical system_type code resolution table
_SCOPE_CODE_MAP: List[Tuple[str, str]] = [
    ("VRF",               "VRF"),
    ("VARIABLE REFRIGERANT", "VRF"),
    ("CHILLER",           "CHILLER"),
    ("CHILLED WATER",     "CHILLER"),
    ("FCU_CW",            "FCU"),
    ("FCU",               "FCU"),
    ("FAN COIL",          "FCU"),
    ("CASSETTE",          "CASSETTE"),
    ("PACKAGED_DX",       "PACKAGED_DX"),
    ("PACKAGED DX",       "PACKAGED_DX"),
    ("PACKAGED UNIT",     "PACKAGED_DX"),
    ("PACKAGED",          "PACKAGED_DX"),
    ("SPLIT_AC",          "SPLIT_AC"),
    ("SPLIT AC",          "SPLIT_AC"),
    ("SPLIT AIR",         "SPLIT_AC"),
    ("SPLIT",             "SPLIT_AC"),
    ("AHU",               "AHU"),
    ("AIR HANDLING",      "AHU"),
]


# ---------------------------------------------------------------------------
# Result dataclass
# ---------------------------------------------------------------------------
@dataclass
class RFQResult:
    """Output produced by RFQGeneratorAgent.run()."""

    xlsx_bytes: bytes = field(default_factory=bytes)
    pdf_bytes: bytes = field(default_factory=bytes)
    rfq_ref: str = ""
    filename_xlsx: str = ""
    filename_pdf: str = ""
    system_code: str = ""
    system_label: str = ""
    selection_basis: str = ""
    confidence_pct: int = 0
    scope_rows: List[ScopeRow] = field(default_factory=list)
    rfq_record: Any = None          # GeneratedRFQ | None
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Main agent class
# ---------------------------------------------------------------------------
class RFQGeneratorAgent:
    """Builds Excel + PDF RFQ documents from a ProcurementRequest.

    Selection modes
    ---------------
    "RECOMMENDED"   -- fetches the latest RecommendationResult.output_payload_json
                       and uses its system_type_code.  Triggered when the user
                       selects "Use Recommendation" in the RFQ generator modal.
    Any other str   -- treated as a direct HVAC system code (e.g. "VRF", "SPLIT_AC").
                       Triggered when the user manually picks a system in the modal.
    """

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------
    @classmethod
    def run(
        cls,
        proc_request: Any,
        *,
        selection_mode: str = "RECOMMENDED",
        qty_overrides: Optional[Dict[int, Any]] = None,
        generated_by: Any = None,
        save_record: bool = False,
    ) -> RFQResult:
        """Generate RFQ Excel + PDF for proc_request.

        Parameters
        ----------
        proc_request    : ProcurementRequest instance.
        selection_mode  : "RECOMMENDED" to use AI recommendation; otherwise an
                          HVAC system code string (e.g. "VRF", "SPLIT_AC").
                          When the form "Use Recommendation" button is clicked,
                          the view passes selection_mode="RECOMMENDED".
        qty_overrides   : Optional dict mapping scope-row index -> custom qty.
        generated_by    : User instance (for audit purposes).
        save_record     : If True, upload blobs + persist GeneratedRFQ DB record.

        Returns
        -------
        RFQResult with all generated content and optional DB record.
        """
        try:
            return cls._generate(
                proc_request,
                selection_mode=selection_mode,
                qty_overrides=qty_overrides or {},
                generated_by=generated_by,
                save_record=save_record,
            )
        except Exception as exc:
            logger.exception("RFQGeneratorAgent.run failed for request pk=%s: %s", proc_request.pk, exc)
            return RFQResult(error=str(exc))

    # ------------------------------------------------------------------
    # Internal generation pipeline
    # ------------------------------------------------------------------
    @classmethod
    def _generate(
        cls,
        proc_request: Any,
        *,
        selection_mode: str,
        qty_overrides: Dict[int, Any],
        generated_by: Any,
        save_record: bool,
    ) -> RFQResult:

        # ----------------------------------------------------------
        # 1. Resolve system code + recommendation metadata
        # ----------------------------------------------------------
        system_code, rationale, confidence_pct, selection_basis = cls._resolve_system(
            proc_request, selection_mode
        )

        # ----------------------------------------------------------
        # 2. Load scope rows (DB-first; fall back to hardcoded table)
        # ----------------------------------------------------------
        scope_rows, system_label, capacity_note = cls._load_scope(
            proc_request, system_code
        )

        # Apply quantity overrides (index-based)
        if qty_overrides:
            scope_rows = [
                (cat, desc, unit, qty_overrides.get(i, qty))
                for i, (cat, desc, unit, qty) in enumerate(scope_rows)
            ]

        # ----------------------------------------------------------
        # 3. Gather request attributes
        # ----------------------------------------------------------
        attrs = cls._gather_attrs(proc_request)
        country      = attrs.get("country", "") or getattr(proc_request, "geography_country", "") or ""
        city         = attrs.get("city", "") or getattr(proc_request, "geography_city", "") or ""
        store_type   = attrs.get("store_type", "")
        area_sqft    = attrs.get("area_sqft", "")
        ceiling_h    = attrs.get("ceiling_height", "")
        ambient      = attrs.get("ambient_temp_max", "")
        humidity     = attrs.get("humidity_level", "")
        cooling_tr   = attrs.get("estimated_cooling_tr", "")
        budget       = attrs.get("budget_level", "")

        today = datetime.date.today().strftime("%d-%b-%Y")
        rfq_ref = f"RFQ-{proc_request.pk:04d}-{datetime.date.today().strftime('%Y%m%d')}"
        capacity_display = f"{cooling_tr} TR" if cooling_tr else capacity_note
        safe_title = cls._safe_title(proc_request)
        filename_xlsx = f"RFQ_{rfq_ref}_{safe_title}.xlsx"
        filename_pdf = f"RFQ_{rfq_ref}_{safe_title}.pdf"

        # ----------------------------------------------------------
        # 4. Build Excel workbook
        # ----------------------------------------------------------
        logger.info(
            "RFQGeneratorAgent: building Excel for request pk=%s system=%s mode=%s",
            proc_request.pk, system_code, selection_mode,
        )
        xlsx_bytes = cls._build_excel(
            proc_request=proc_request,
            rfq_ref=rfq_ref,
            today=today,
            system_label=system_label,
            system_code=system_code,
            selection_basis=selection_basis,
            confidence_pct=confidence_pct,
            rationale=rationale,
            capacity_display=capacity_display,
            scope_rows=scope_rows,
            country=country,
            city=city,
            store_type=store_type,
            area_sqft=area_sqft,
            ceiling_h=ceiling_h,
            ambient=ambient,
            humidity=humidity,
            budget=budget,
        )

        # ----------------------------------------------------------
        # 5. Build PDF
        # ----------------------------------------------------------
        pdf_bytes = cls._build_pdf(
            proc_request=proc_request,
            rfq_ref=rfq_ref,
            today=today,
            system_label=system_label,
            selection_basis=selection_basis,
            confidence_pct=confidence_pct,
            rationale=rationale,
            capacity_display=capacity_display,
            scope_rows=scope_rows,
            country=country,
            city=city,
            store_type=store_type,
            area_sqft=area_sqft,
            ceiling_h=ceiling_h,
            ambient=ambient,
            humidity=humidity,
            budget=budget,
        )

        # ----------------------------------------------------------
        # 6. Blob upload + DB persistence (only when save_record=True)
        # ----------------------------------------------------------
        rfq_record = None
        if save_record:
            rfq_record = cls._persist(
                proc_request=proc_request,
                rfq_ref=rfq_ref,
                safe_title=safe_title,
                system_code=system_code,
                system_label=system_label,
                qty_overrides=qty_overrides,
                xlsx_bytes=xlsx_bytes,
                pdf_bytes=pdf_bytes,
                generated_by=generated_by,
            )

        return RFQResult(
            xlsx_bytes=xlsx_bytes,
            pdf_bytes=pdf_bytes,
            rfq_ref=rfq_ref,
            filename_xlsx=filename_xlsx,
            filename_pdf=filename_pdf,
            system_code=system_code,
            system_label=system_label,
            selection_basis=selection_basis,
            confidence_pct=confidence_pct,
            scope_rows=scope_rows,
            rfq_record=rfq_record,
        )

    # ------------------------------------------------------------------
    # 1. System resolution
    # ------------------------------------------------------------------
    @classmethod
    def _resolve_system(
        cls,
        proc_request: Any,
        selection_mode: str,
    ) -> Tuple[str, str, int, str]:
        """Return (system_code, rationale, confidence_pct, selection_basis).

        When selection_mode == "RECOMMENDED" (triggered by "Use Recommendation"
        button in the RFQ modal form), the latest RecommendationResult is queried
        and its system_type_code is used to drive the document.
        """
        from apps.procurement.models import RecommendationResult

        param = selection_mode.strip().upper()

        if param == "RECOMMENDED":
            # ----------------------------------------------------------
            # Form "Use Recommendation" path:
            # Look up the most recent RecommendationResult for this request
            # and extract the AI-recommended system type code.
            # ----------------------------------------------------------
            rec = (
                RecommendationResult.objects
                .filter(run__request=proc_request)
                .order_by("-created_at")
                .first()
            )
            if rec:
                _raw = (
                    (rec.output_payload_json or {}).get("system_type_code", "")
                    or str(rec.recommended_option or "").split("(")[0].strip()
                )
                system_code = cls._normalize_system_code(_raw)
                rationale = rec.reasoning_summary or (
                    f"Auto-selected from latest recommendation ({system_code})."
                )
                confidence_pct = round(float(rec.confidence_score or 0) * 100)
                logger.info(
                    "RFQGeneratorAgent: RECOMMENDED mode -- system_code=%s confidence=%s%% "
                    "(RecommendationResult pk=%s)",
                    system_code, confidence_pct, rec.pk,
                )
            else:
                raise ValueError(
                    "RFQ generation requires a completed recommendation when selection_mode="
                    "'RECOMMENDED'. No RecommendationResult was found for request "
                    f"pk={proc_request.pk}."
                )
            selection_basis = "AI / Rules Engine Recommendation"
        else:
            # ----------------------------------------------------------
            # Manual selection path (user picked a specific system code
            # in the Step 1 dropdown of the RFQ modal).
            # ----------------------------------------------------------
            system_code = cls._normalize_system_code(param)
            rationale = f"Manual system selection: {system_code}."
            confidence_pct = 0
            selection_basis = "Manual Selection"
            logger.info(
                "RFQGeneratorAgent: MANUAL mode -- system_code=%s (raw param=%s)",
                system_code, selection_mode,
            )

        return system_code, rationale, confidence_pct, selection_basis

    # ------------------------------------------------------------------
    # 2. Scope loading
    # ------------------------------------------------------------------
    @classmethod
    def _load_scope(
        cls,
        proc_request: Any,
        system_code: str,
    ) -> Tuple[List[ScopeRow], str, str]:
        """Return (scope_rows, system_label, capacity_note).

        DB HVACServiceScope rows are mandatory.
        """
        from apps.procurement.models import HVACServiceScope

        db_scope = HVACServiceScope.objects.filter(system_type=system_code, is_active=True).first()
        if not db_scope:
            db_scope = HVACServiceScope.objects.filter(
                system_type__iexact=system_code, is_active=True
            ).first()

        if db_scope:
            system_label = db_scope.display_name or _SYSTEM_LABELS.get(system_code, system_code)
            raw_rows: List[ScopeRow] = []
            for cat, field_text in [
                ("Equipment",      db_scope.equipment_scope),
                ("Installation",   db_scope.installation_services),
                ("Piping/Ducting", db_scope.piping_ducting),
                ("Electrical",     db_scope.electrical_works),
                ("Controls",       db_scope.controls_accessories),
                ("Testing",        db_scope.testing_commissioning),
            ]:
                added = 0
                for line in (field_text or "").splitlines():
                    line = line.strip().lstrip("-*. ").strip()
                    if line:
                        raw_rows.append((cat, line, "LS", 1))
                        added += 1
                if added == 0:
                    raise ValueError(
                        "RFQ scope is incomplete in HVACServiceScope for system "
                        f"'{system_code}' (missing entries in category '{cat}')."
                    )
            capacity_note = "As per heat load calculation (TR)"
            logger.info(
                "RFQGeneratorAgent: loaded %d scope rows from DB (system=%s request pk=%s)",
                len(raw_rows), system_code, proc_request.pk,
            )
            return raw_rows, system_label, capacity_note
        raise ValueError(
            "RFQ scope configuration is missing. No active HVACServiceScope found for "
            f"system '{system_code}' (request pk={proc_request.pk})."
        )

    # ------------------------------------------------------------------
    # 3. Attribute gathering
    # ------------------------------------------------------------------
    @staticmethod
    def _gather_attrs(proc_request: Any) -> Dict[str, str]:
        """Return a flat dict of attribute_code -> value_text for the request."""
        try:
            return {a.attribute_code: a.value_text for a in proc_request.attributes.all()}
        except Exception:
            return {}

    # ------------------------------------------------------------------
    # 4. Excel builder
    # ------------------------------------------------------------------
    @classmethod
    def _build_excel(
        cls,
        *,
        proc_request: Any,
        rfq_ref: str,
        today: str,
        system_label: str,
        system_code: str,
        selection_basis: str,
        confidence_pct: int,
        rationale: str,
        capacity_display: str,
        scope_rows: List[ScopeRow],
        country: str,
        city: str,
        store_type: str,
        area_sqft: str,
        ceiling_h: str,
        ambient: str,
        humidity: str,
        budget: str,
    ) -> bytes:
        """Build the RFQ Excel workbook and return raw bytes."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RFQ"
        ws.sheet_view.showGridLines = False

        # Column widths: A(idx)=6, B(param)=28, C(value/desc)=48, D(unit)=14, E(qty)=10
        for i, w in zip(range(1, 6), [6, 28, 48, 14, 10]):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ---- Style helpers ----
        def _thin():
            s = Side(style="thin", color="BDBDBD")
            return Border(left=s, right=s, top=s, bottom=s)

        def _fill(hex_val):
            return PatternFill("solid", fgColor=hex_val)

        NAVY  = "1A3C5E"
        LTBLUE = "EAF2FB"
        GREY   = "F5F5F5"
        WHITE  = "FFFFFF"
        SECBG  = "D6E4F0"

        def _merge_write(row, col_start, col_end, value, font=None, fill=None,
                         align=None, height=None, border=True):
            ws.merge_cells(start_row=row, start_column=col_start,
                           end_row=row, end_column=col_end)
            c = ws.cell(row=row, column=col_start, value=value)
            if font:  c.font = font
            if fill:  c.fill = fill
            if align: c.alignment = align
            if border:
                for col in range(col_start, col_end + 1):
                    ws.cell(row=row, column=col).border = _thin()
            if height:
                ws.row_dimensions[row].height = height

        def _kv(row, label, value, label_fill=GREY):
            lc = ws.cell(row=row, column=2, value=label)
            lc.font = Font(bold=True, size=10)
            lc.fill = _fill(label_fill)
            lc.border = _thin()
            lc.alignment = Alignment(vertical="center")
            vc = ws.cell(row=row, column=3, value=value or "--")
            vc.font = Font(size=10)
            vc.fill = _fill(WHITE)
            vc.border = _thin()
            vc.alignment = Alignment(vertical="center", wrap_text=True)
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
            for col in range(3, 6):
                ws.cell(row=row, column=col).border = _thin()
            ws.row_dimensions[row].height = 18

        def _section_header(row, number, title):
            ws.merge_cells(start_row=row, start_column=2,
                           end_row=row, end_column=5)
            c = ws.cell(row=row, column=2, value=f"{number}. {title}")
            c.font = Font(bold=True, size=11, color=NAVY)
            c.fill = _fill(SECBG)
            c.alignment = Alignment(vertical="center")
            c.border = _thin()
            for col in range(2, 6):
                ws.cell(row=row, column=col).border = _thin()
            ws.row_dimensions[row].height = 20

        def _table_header(row, cols):
            for ci, label in cols:
                c = ws.cell(row=row, column=ci, value=label)
                c.font = Font(bold=True, size=10, color=NAVY)
                c.fill = _fill(LTBLUE)
                c.border = _thin()
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 18

        # ============================== ROW 1: TITLE ==============================
        _merge_write(
            1, 1, 5,
            "REQUEST FOR QUOTATION (RFQ) - HVAC WORKS",
            font=Font(bold=True, size=16, color="FFFFFF"),
            fill=_fill(NAVY),
            align=Alignment(horizontal="center", vertical="center"),
            height=36, border=False,
        )

        # Row 2: Ref + Date
        ws.cell(row=2, column=2, value=f"RFQ Ref: {rfq_ref}").font = Font(bold=True, size=9, color="555555")
        ws.merge_cells("B2:C2")
        date_c = ws.cell(row=2, column=4, value=f"Date: {today}")
        date_c.font = Font(size=9, color="555555")
        date_c.alignment = Alignment(horizontal="right")
        ws.merge_cells("D2:E2")
        ws.row_dimensions[2].height = 14

        # Row 3: AI Recommendation badge (shown when selection_basis is AI)
        if selection_basis == "AI / Rules Engine Recommendation":
            ws.merge_cells("B3:E3")
            badge = ws.cell(
                row=3, column=2,
                value=f"Generated using AI Recommendation  |  System: {system_code}  |  Confidence: {confidence_pct}%",
            )
            badge.font = Font(size=9, italic=True, color="FFFFFF")
            badge.fill = _fill("2E7D32")   # dark green
            badge.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[3].height = 14
            intro_row = 5
        else:
            intro_row = 4

        # ========================== INTRO PARAGRAPH ================================
        intro = (
            f"We invite your quotation for the Supply, Installation, Testing, and Commissioning (SITC) "
            f"of a {system_label} HVAC system for the store described below. "
            f"Please submit a detailed, itemised quotation covering all scope items listed in Section 3."
        )
        _merge_write(
            intro_row, 2, 5, intro,
            font=Font(size=10),
            fill=_fill(WHITE),
            align=Alignment(wrap_text=True, vertical="top"),
            height=40, border=False,
        )
        ws.row_dimensions[intro_row].height = 44

        # ========================== SECTION 1: STORE DETAILS ======================
        r = intro_row + 2
        _section_header(r, 1, "Store Details"); r += 1
        _table_header(r, [(2, "Parameter"), (3, "Value")]); r += 1
        store_rows = [
            ("Country",          country),
            ("City",             city),
            ("Store Type",       store_type),
            ("Area",             f"{area_sqft} sq ft" if area_sqft else ""),
            ("Ceiling Height",   f"{ceiling_h} ft" if ceiling_h else ""),
            ("Max Temperature",  f"{ambient} deg C" if ambient else ""),
            ("Humidity",         humidity),
            ("Budget Level",     budget),
        ]
        for param, val in store_rows:
            if not val:
                continue
            _kv(r, param, val); r += 1

        # ========================= SECTION 2: HVAC SYSTEM =========================
        r += 1
        _section_header(r, 2, "Recommended HVAC System"); r += 1
        _table_header(r, [(2, "Field"), (3, "Value")]); r += 1
        _kv(r, "System Type",     system_label); r += 1
        _kv(r, "Capacity",        capacity_display); r += 1
        _kv(r, "Selection Basis", selection_basis); r += 1
        if confidence_pct:
            _kv(r, "Confidence", f"{confidence_pct}%"); r += 1

        # Rationale row (taller)
        lc2 = ws.cell(row=r, column=2, value="Reason / Rationale")
        lc2.font = Font(bold=True, size=10)
        lc2.fill = _fill(GREY)
        lc2.border = _thin()
        lc2.alignment = Alignment(vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        vc2 = ws.cell(row=r, column=3, value=rationale)
        vc2.font = Font(size=10, italic=True)
        vc2.fill = _fill(WHITE)
        vc2.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(2, 6):
            ws.cell(row=r, column=col).border = _thin()
        ws.row_dimensions[r].height = 50
        r += 1

        # ========================== SECTION 3: SCOPE OF WORK ======================
        r += 1
        _section_header(r, 3, "Scope of Work"); r += 1
        scope_hdr_cols = [(1, "S.No"), (2, "Category"), (3, "Description"), (4, "Unit"), (5, "Qty")]
        for ci, label in scope_hdr_cols:
            c = ws.cell(row=r, column=ci, value=label)
            c.font = Font(bold=True, size=10, color=NAVY)
            c.fill = _fill(LTBLUE)
            c.border = _thin()
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1

        for sno, (cat, desc, unit, qty) in enumerate(scope_rows, 1):
            ws.cell(row=r, column=1, value=sno).font = Font(size=10)
            ws.cell(row=r, column=1).fill = _fill(GREY)
            ws.cell(row=r, column=1).border = _thin()
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row=r, column=2, value=cat).font = Font(bold=True, size=10)
            ws.cell(row=r, column=2).fill = _fill(GREY)
            ws.cell(row=r, column=2).border = _thin()
            ws.cell(row=r, column=2).alignment = Alignment(vertical="center")

            ws.cell(row=r, column=3, value=desc).font = Font(size=10)
            ws.cell(row=r, column=3).fill = _fill(WHITE)
            ws.cell(row=r, column=3).border = _thin()
            ws.cell(row=r, column=3).alignment = Alignment(vertical="center", wrap_text=True)

            ws.cell(row=r, column=4, value=unit).font = Font(size=10)
            ws.cell(row=r, column=4).fill = _fill(WHITE)
            ws.cell(row=r, column=4).border = _thin()
            ws.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row=r, column=5, value=qty).font = Font(size=10)
            ws.cell(row=r, column=5).fill = _fill(WHITE)
            ws.cell(row=r, column=5).border = _thin()
            ws.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[r].height = 20
            r += 1

        # ========================== SECTION 4: COMMERCIAL TERMS ===================
        r += 1
        _section_header(r, 4, "Commercial Terms"); r += 1
        _table_header(r, [(2, "Term"), (3, "Details")]); r += 1
        commercial_rows = [
            ("Quotation Validity",    "90 days from submission date"),
            ("Submission Deadline",   "As advised by issuing party"),
            ("Delivery / Completion", "Vendor to specify in quotation"),
            ("Payment Terms",         "As per standard purchase order terms"),
            ("Currency",              f"{country} local currency or as agreed"),
            ("Pricing Basis",         "Lump sum (supply + installation + commissioning)"),
            ("Warranty",              "Minimum 2 years on equipment and workmanship"),
            ("Compliance",            "ASHRAE 90.1, SASO, and applicable local building code"),
        ]
        for term, detail in commercial_rows:
            _kv(r, term, detail); r += 1

        # ========================== SECTION 5: VENDOR PRICING RESPONSE ============
        r += 1
        _section_header(r, 5, "Vendor Pricing Response (To be filled by Vendor)"); r += 1
        vnd_cols = [
            (1, "S.No"), (2, "Category"),
            (3, "Description / Proposed Model"),
            (4, "Unit Price"), (5, "Total"),
        ]
        for ci, label in vnd_cols:
            c = ws.cell(row=r, column=ci, value=label)
            c.font = Font(bold=True, size=10, color=NAVY)
            c.fill = _fill(LTBLUE)
            c.border = _thin()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 18
        r += 1
        for idx in range(1, len(scope_rows) + 1):
            for ci in range(1, 6):
                c = ws.cell(row=r, column=ci, value=idx if ci == 1 else "")
                c.border = _thin()
                c.fill = _fill(WHITE)
                c.font = Font(size=10)
                if ci in (4, 5):
                    c.alignment = Alignment(horizontal="right")
            ws.row_dimensions[r].height = 18
            r += 1

        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        tc = ws.cell(row=r, column=3, value="GRAND TOTAL (Excl. VAT)")
        tc.font = Font(bold=True, size=10)
        tc.fill = _fill(SECBG)
        tc.border = _thin()
        tc.alignment = Alignment(horizontal="right", vertical="center")
        gt = ws.cell(row=r, column=5, value="")
        gt.border = _thin()
        gt.fill = _fill(SECBG)
        gt.font = Font(bold=True, size=10)
        ws.row_dimensions[r].height = 18
        r += 2

        # Disclaimer
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        dc = ws.cell(
            row=r, column=1,
            value=(
                "Note: This RFQ is system-generated. All specifications are indicative. "
                "A qualified HVAC engineer must review and confirm final scope before award."
            ),
        )
        dc.font = Font(size=8, italic=True, color="888888")
        dc.alignment = Alignment(wrap_text=True, horizontal="center")
        dc.fill = _fill("FFFDE7")
        ws.row_dimensions[r].height = 24

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # 5. PDF builder
    # ------------------------------------------------------------------
    @classmethod
    def _build_pdf(
        cls,
        *,
        proc_request: Any,
        rfq_ref: str,
        today: str,
        system_label: str,
        selection_basis: str,
        confidence_pct: int,
        rationale: str,
        capacity_display: str,
        scope_rows: List[ScopeRow],
        country: str,
        city: str,
        store_type: str,
        area_sqft: str,
        ceiling_h: str,
        ambient: str,
        humidity: str,
        budget: str,
    ) -> bytes:
        """Build the RFQ PDF document and return raw bytes.

        Returns empty bytes if reportlab is not installed.
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        except ImportError:
            logger.warning("RFQGeneratorAgent: reportlab not installed -- skipping PDF generation")
            return b""

        try:
            pdf_buf = io.BytesIO()
            doc = SimpleDocTemplate(
                pdf_buf, pagesize=A4,
                leftMargin=18 * mm, rightMargin=18 * mm,
                topMargin=18 * mm, bottomMargin=18 * mm,
            )
            styles = getSampleStyleSheet()
            NAVY_RL   = colors.HexColor("#1A3C5E")
            LTBLUE_RL = colors.HexColor("#EAF2FB")
            SECBG_RL  = colors.HexColor("#D6E4F0")
            GREY_RL   = colors.HexColor("#F5F5F5")
            GREEN_RL  = colors.HexColor("#2E7D32")

            title_style = ParagraphStyle(
                "rfqTitle", parent=styles["Title"],
                fontSize=14, textColor=colors.white,
                backColor=NAVY_RL, spaceAfter=4, spaceBefore=0,
                alignment=1,
            )
            sec_style = ParagraphStyle(
                "rfqSec", parent=styles["Normal"],
                fontSize=11, textColor=NAVY_RL, fontName="Helvetica-Bold",
                spaceAfter=2, spaceBefore=6,
            )
            normal_sm = ParagraphStyle(
                "rfqNorm", parent=styles["Normal"],
                fontSize=9, spaceAfter=1,
            )
            italic_sm = ParagraphStyle(
                "rfqItalic", parent=styles["Normal"],
                fontSize=9, fontName="Helvetica-Oblique", spaceAfter=4,
            )
            badge_style = ParagraphStyle(
                "rfqBadge", parent=styles["Normal"],
                fontSize=8, textColor=colors.white,
                backColor=GREEN_RL, spaceAfter=3, alignment=1,
            )

            story = []

            # Title
            story.append(Paragraph("REQUEST FOR QUOTATION (RFQ) -- HVAC WORKS", title_style))
            story.append(Spacer(1, 4 * mm))

            # Ref + Date
            story.append(Paragraph(
                f"<b>RFQ Ref:</b> {rfq_ref} &nbsp;&nbsp;&nbsp; <b>Date:</b> {today}",
                normal_sm,
            ))

            # AI Recommendation badge
            if selection_basis == "AI / Rules Engine Recommendation":
                story.append(Paragraph(
                    f"<b>AI Recommendation</b> | Confidence: {confidence_pct}%",
                    badge_style,
                ))

            story.append(Spacer(1, 3 * mm))

            # Intro
            story.append(Paragraph(
                f"We invite your quotation for the Supply, Installation, Testing, and "
                f"Commissioning (SITC) of a <b>{system_label}</b> HVAC system for the store "
                f"described below. Please submit a detailed, itemised quotation covering all "
                f"scope items in Section 3.",
                italic_sm,
            ))
            story.append(Spacer(1, 2 * mm))

            def _kv_table(rows, col_widths=(55 * mm, 115 * mm)):
                data = [
                    [Paragraph(f"<b>{k}</b>", normal_sm), Paragraph(str(v or "--"), normal_sm)]
                    for k, v in rows
                ]
                t = Table(data, colWidths=col_widths)
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (0, -1), GREY_RL),
                    ("BACKGROUND", (1, 0), (1, -1), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]))
                return t

            # Section 1: Store Details
            story.append(Paragraph("1. Store Details", sec_style))
            store_kv = [(p, v) for p, v in [
                ("Country", country),
                ("City", city),
                ("Store Type", store_type),
                ("Area", f"{area_sqft} sq ft" if area_sqft else ""),
                ("Ceiling Height", f"{ceiling_h} ft" if ceiling_h else ""),
                ("Max Temperature", f"{ambient} deg C" if ambient else ""),
                ("Humidity", humidity),
                ("Budget Level", budget),
            ] if v]
            story.append(_kv_table(store_kv))
            story.append(Spacer(1, 3 * mm))

            # Section 2: HVAC System
            story.append(Paragraph("2. Recommended HVAC System", sec_style))
            sys_kv = [
                ("System Type", system_label),
                ("Capacity", capacity_display),
                ("Selection Basis", selection_basis),
            ]
            if confidence_pct:
                sys_kv.append(("Confidence", f"{confidence_pct}%"))
            sys_kv.append(("Reason / Rationale", rationale))
            story.append(_kv_table(sys_kv))
            story.append(Spacer(1, 3 * mm))

            # Section 3: Scope of Work
            story.append(Paragraph("3. Scope of Work", sec_style))
            hdr = [
                Paragraph("<b>S.No</b>", normal_sm),
                Paragraph("<b>Category</b>", normal_sm),
                Paragraph("<b>Description</b>", normal_sm),
                Paragraph("<b>Unit</b>", normal_sm),
                Paragraph("<b>Qty</b>", normal_sm),
            ]
            scope_pdf_rows = [hdr] + [
                [
                    Paragraph(str(i), normal_sm),
                    Paragraph(cat, normal_sm),
                    Paragraph(desc, normal_sm),
                    Paragraph(unit, normal_sm),
                    Paragraph(str(qty) if qty != "" else "", normal_sm),
                ]
                for i, (cat, desc, unit, qty) in enumerate(scope_rows, 1)
            ]
            scope_t = Table(scope_pdf_rows, colWidths=[10 * mm, 28 * mm, 80 * mm, 20 * mm, 14 * mm])
            scope_t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), LTBLUE_RL),
                ("BACKGROUND", (0, 1), (1, -1), GREY_RL),
                ("BACKGROUND", (2, 1), (-1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]))
            story.append(scope_t)
            story.append(Spacer(1, 3 * mm))

            # Section 4: Commercial Terms
            story.append(Paragraph("4. Commercial Terms", sec_style))
            comm_kv = [
                ("Quotation Validity",    "90 days from submission date"),
                ("Delivery / Completion", "Vendor to specify in quotation"),
                ("Payment Terms",         "As per standard purchase order terms"),
                ("Currency",              f"{country} local currency or as agreed"),
                ("Pricing Basis",         "Lump sum (supply + installation + commissioning)"),
                ("Warranty",              "Minimum 2 years on equipment and workmanship"),
                ("Compliance",            "ASHRAE 90.1, SASO, and applicable local building code"),
            ]
            story.append(_kv_table(comm_kv))
            story.append(Spacer(1, 3 * mm))

            # Section 5: Vendor Pricing Response
            story.append(Paragraph("5. Vendor Pricing Response (To be filled by Vendor)", sec_style))
            vnd_hdr = [
                Paragraph("<b>S.No</b>", normal_sm),
                Paragraph("<b>Category</b>", normal_sm),
                Paragraph("<b>Description / Proposed Model</b>", normal_sm),
                Paragraph("<b>Unit Price</b>", normal_sm),
                Paragraph("<b>Total</b>", normal_sm),
            ]
            vnd_rows = [vnd_hdr] + [
                [
                    Paragraph(str(i), normal_sm),
                    Paragraph("", normal_sm),
                    Paragraph("", normal_sm),
                    Paragraph("", normal_sm),
                    Paragraph("", normal_sm),
                ]
                for i in range(1, len(scope_rows) + 1)
            ] + [[
                Paragraph("", normal_sm),
                Paragraph("", normal_sm),
                Paragraph("<b>GRAND TOTAL (Excl. VAT)</b>", normal_sm),
                Paragraph("", normal_sm),
                Paragraph("", normal_sm),
            ]]
            vnd_t = Table(vnd_rows, colWidths=[10 * mm, 28 * mm, 80 * mm, 20 * mm, 14 * mm])
            vnd_t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), LTBLUE_RL),
                ("BACKGROUND", (-3, -1), (-1, -1), SECBG_RL),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(vnd_t)
            story.append(Spacer(1, 4 * mm))

            # Disclaimer
            story.append(Paragraph(
                "<i>Note: This RFQ is system-generated. All specifications are indicative. "
                "A qualified HVAC engineer must review and confirm the final scope before award.</i>",
                ParagraphStyle(
                    "disc", parent=styles["Normal"],
                    fontSize=7, textColor=colors.HexColor("#888888"),
                ),
            ))

            doc.build(story)
            pdf_buf.seek(0)
            return pdf_buf.getvalue()

        except Exception as exc:
            logger.warning("RFQGeneratorAgent PDF build failed: %s", exc)
            return b""

    # ------------------------------------------------------------------
    # 6. Blob upload + DB persistence
    # ------------------------------------------------------------------
    @classmethod
    def _persist(
        cls,
        *,
        proc_request: Any,
        rfq_ref: str,
        safe_title: str,
        system_code: str,
        system_label: str,
        qty_overrides: Dict[int, Any],
        xlsx_bytes: bytes,
        pdf_bytes: bytes,
        generated_by: Any,
    ) -> Any:
        """Upload xlsx + pdf to Azure Blob and save a GeneratedRFQ DB record.

        Returns the created GeneratedRFQ instance.
        """
        from apps.procurement.models import GeneratedRFQ

        xlsx_blob_path = ""
        pdf_blob_path = ""
        _date_str = datetime.date.today().strftime("%Y%m%d")
        _base_name = f"RFQ-{proc_request.pk:04d}-{_date_str}_{safe_title}"
        _folder = f"rfq/{safe_title}"

        try:
            from apps.documents.blob_service import upload_to_blob, is_blob_storage_enabled
            if is_blob_storage_enabled():
                if xlsx_bytes:
                    _xblob = f"{_folder}/{_base_name}.xlsx"
                    upload_to_blob(
                        io.BytesIO(xlsx_bytes),
                        _xblob,
                        content_type=(
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        ),
                    )
                    xlsx_blob_path = _xblob
                if pdf_bytes:
                    _pblob = f"{_folder}/{_base_name}.pdf"
                    upload_to_blob(
                        io.BytesIO(pdf_bytes),
                        _pblob,
                        content_type="application/pdf",
                    )
                    pdf_blob_path = _pblob
                logger.info(
                    "RFQGeneratorAgent: uploaded to blob -- xlsx=%s pdf=%s",
                    xlsx_blob_path or "(none)", pdf_blob_path or "(none)",
                )
        except Exception as blob_exc:
            logger.warning("RFQGeneratorAgent: blob upload failed: %s", blob_exc)

        rfq_record = GeneratedRFQ.objects.create(
            request=proc_request,
            rfq_ref=rfq_ref,
            system_code=system_code,
            system_label=system_label,
            qty_json=qty_overrides,
            xlsx_blob_path=xlsx_blob_path,
            pdf_blob_path=pdf_blob_path,
            generated_by=generated_by if (generated_by and generated_by.is_authenticated) else None,
        )
        logger.info(
            "RFQGeneratorAgent: GeneratedRFQ pk=%s saved for request pk=%s",
            rfq_record.pk, proc_request.pk,
        )
        return rfq_record

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _normalize_system_code(raw: str) -> str:
        """Resolve any text representation of an HVAC system name to its
        canonical system_type key (e.g. 'Split Air Conditioning' -> 'SPLIT_AC').

        Falls back to returning the input uppercased when no known keyword matches.
        """
        if not raw:
            raise ValueError("HVAC system code is required for RFQ generation.")
        _u = raw.strip().upper()
        # 1. Exact match
        for _kw, _code in _SCOPE_CODE_MAP:
            if _u == _kw:
                return _code
        # 2. Startswith
        for _kw, _code in _SCOPE_CODE_MAP:
            if _u.startswith(_kw):
                return _code
        # 3. Contains
        for _kw, _code in _SCOPE_CODE_MAP:
            if _kw in _u:
                return _code
        return _u   # valid DB key already (caller's responsibility)

    @staticmethod
    def _safe_title(proc_request: Any) -> str:
        """Return a filesystem-safe title slug for the request (max 30 chars)."""
        raw = proc_request.title or "Request"
        return (
            "".join(c for c in raw if c.isalnum() or c in " _-")[:30]
            .strip()
            .replace(" ", "_")
        )
