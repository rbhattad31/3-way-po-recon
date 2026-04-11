"""Template views for the Procurement Intelligence UI."""
from __future__ import annotations

import json
import logging
import re
import requests
from urllib.parse import urlparse

from django.conf import settings

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST, require_http_methods
from django.db.models import Count, OuterRef, Q, Subquery
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render

from apps.auditlog.services import AuditService
from apps.core.permissions import permission_required_code
from apps.core.enums import (
    AnalysisRunType,
    ProcurementRequestStatus,
    ProcurementRequestType,
    ExternalSourceClass,
)
from apps.procurement.models import (
    AnalysisRun,
    BenchmarkResult,
    ComplianceResult,
    ExternalSourceRegistry,
    HVACStoreProfile,
    ProcurementRequest,
    ProcurementRequestAttribute,
    RecommendationResult,
    SupplierQuotation,
    ValidationResult,
)

from apps.procurement.agents.reason_summary_agent import ReasonSummaryAgent
from apps.agents.services.llm_client import LLMClient, LLMMessage

logger = logging.getLogger(__name__)


HVAC_MANDATORY_FIELDS = [
    ("f_country", "Country"),
    ("f_city", "City"),
    ("f_store_type", "Store Type"),
    ("f_area_sqft", "Area (sq ft)"),
    ("f_ambient_temp_max", "Ambient Temp Max (C)"),
    ("f_budget_level", "Budget"),
    ("f_energy_efficiency_priority", "Energy Priority"),
]


def _safe_float(value):
    try:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        return float(text)
    except (TypeError, ValueError):
        return None


def _build_hvac_store_profile_defaults(post_data) -> dict:
    return {
        "brand": (post_data.get("f_brand") or "").strip(),
        "country": (post_data.get("f_country") or "").strip(),
        "city": (post_data.get("f_city") or "").strip(),
        "store_type": (post_data.get("f_store_type") or "").strip(),
        "store_format": (post_data.get("f_store_format") or "").strip(),
        "area_sqft": _safe_float(post_data.get("f_area_sqft")),
        "ceiling_height_ft": _safe_float(post_data.get("f_ceiling_height_ft")),
        "operating_hours": (post_data.get("f_operating_hours") or "").strip(),
        "footfall_category": (post_data.get("f_footfall_category") or "").strip(),
        "ambient_temp_max": _safe_float(post_data.get("f_ambient_temp_max")),
        "humidity_level": (post_data.get("f_humidity_level") or "").strip(),
        "dust_exposure": (post_data.get("f_dust_exposure") or "").strip(),
        "heat_load_category": (post_data.get("f_heat_load_category") or "").strip(),
        "fresh_air_requirement": (post_data.get("f_fresh_air_requirement") or "").strip(),
        "landlord_constraints": (post_data.get("f_landlord_constraints") or "").strip(),
        "existing_hvac_type": (post_data.get("f_existing_hvac_type") or "").strip(),
        "budget_level": (post_data.get("f_budget_level") or "").strip(),
        "energy_efficiency_priority": (post_data.get("f_energy_efficiency_priority") or "").strip(),
    }


def _hvac_create_context() -> dict:
    return {
        "type_choices": ProcurementRequestType.choices,
        "priority_choices": [
            ("LOW", "Low"),
            ("MEDIUM", "Medium"),
            ("HIGH", "High"),
            ("CRITICAL", "Critical"),
        ],
    }


# ---------------------------------------------------------------------------
# 0. Procurement Home
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view")
def procurement_home(request):
    """HVAC Procurement Intelligence home page."""
    recent_requests = (
        ProcurementRequest.objects
        .select_related("created_by")
        .filter(request_type="HVAC")
        .order_by("-created_at")[:6]
    )
    return render(request, "procurement/home.html", {
        "recent_requests": recent_requests,
    })


# ---------------------------------------------------------------------------
# 1. Request List
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view")
def request_list(request):
    """List all procurement requests with filters."""
    # Subquery: latest recommendation_option for each request
    latest_recommendation_sq = Subquery(
        RecommendationResult.objects.filter(
            run__request=OuterRef("pk"),
        )
        .order_by("-created_at")
        .values("recommended_option")[:1]
    )
    latest_confidence_sq = Subquery(
        RecommendationResult.objects.filter(
            run__request=OuterRef("pk"),
        )
        .order_by("-created_at")
        .values("confidence_score")[:1]
    )

    qs = ProcurementRequest.objects.select_related("created_by", "assigned_to").annotate(
        attribute_count=Count("attributes"),
        quotation_count=Count("quotations"),
        run_count=Count("analysis_runs"),
        latest_recommended_option=latest_recommendation_sq,
        latest_confidence_score=latest_confidence_sq,
    )

    # Filters
    status_filter = request.GET.get("status")
    type_filter = request.GET.get("request_type")
    domain_filter = request.GET.get("domain_code")
    search = request.GET.get("q")

    if status_filter:
        qs = qs.filter(status=status_filter)
    if type_filter:
        qs = qs.filter(request_type=type_filter)
    if domain_filter:
        qs = qs.filter(domain_code=domain_filter)
    if search:
        qs = qs.filter(Q(title__icontains=search) | Q(description__icontains=search))

    qs = qs.order_by("-created_at")
    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get("page"))

    # Distinct domain codes for filter dropdown
    domains = (
        ProcurementRequest.objects.values_list("domain_code", flat=True)
        .distinct()
        .order_by("domain_code")
    )

    return render(request, "procurement/request_list.html", {
        "page_obj": page,
        "status_choices": ProcurementRequestStatus.choices,
        "type_choices": ProcurementRequestType.choices,
        "domains": domains,
        "current_status": status_filter or "",
        "current_type": type_filter or "",
        "current_domain": domain_filter or "",
        "search_query": search or "",
    })


# ---------------------------------------------------------------------------
# 2. Create Request
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.create")
def request_create(request):
    """Create a new procurement request."""
    if request.method == "POST":
        from apps.procurement.services.request_service import ProcurementRequestService

        attrs_data = []
        # Collect dynamic attributes from POST
        attr_codes = request.POST.getlist("attr_code[]")
        attr_labels = request.POST.getlist("attr_label[]")
        attr_values = request.POST.getlist("attr_value[]")
        attr_types = request.POST.getlist("attr_type[]")
        for i, code in enumerate(attr_codes):
            if code.strip():
                attrs_data.append({
                    "attribute_code": code.strip(),
                    "attribute_label": attr_labels[i] if i < len(attr_labels) else code,
                    "data_type": attr_types[i] if i < len(attr_types) else "TEXT",
                    "value_text": attr_values[i] if i < len(attr_values) else "",
                    "is_required": False,
                })

        try:
            proc_request = ProcurementRequestService.create_request(
                title=request.POST.get("title", ""),
                description=request.POST.get("description", ""),
                domain_code=request.POST.get("domain_code", ""),
                schema_code=request.POST.get("schema_code", ""),
                request_type=request.POST.get("request_type", ProcurementRequestType.RECOMMENDATION),
                priority=request.POST.get("priority", "MEDIUM"),
                geography_country=request.POST.get("geography_country", ""),
                geography_city=request.POST.get("geography_city", ""),
                currency=request.POST.get("currency", "USD"),
                created_by=request.user,
                attributes=attrs_data if attrs_data else None,
            )
            messages.success(request, f"Procurement request '{proc_request.title}' created successfully.")
            return redirect("procurement:request_workspace", pk=proc_request.pk)
        except Exception as exc:
            messages.error(request, f"Failed to create request: {exc}")

    return render(request, "procurement/request_create.html", {
        "type_choices": ProcurementRequestType.choices,
        "priority_choices": [("LOW", "Low"), ("MEDIUM", "Medium"), ("HIGH", "High"), ("CRITICAL", "Critical")],
    })


# ---------------------------------------------------------------------------
# 2b. Create HVAC Request (Landmark Group HVAC-specific form)
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.create")
def hvac_create(request):
    """Create a new HVAC procurement request using the Landmark Group HVAC form."""
    if request.method == "POST":
        from apps.procurement.services.request_service import ProcurementRequestService
        from apps.procurement.services.analysis_run_service import AnalysisRunService
        from apps.procurement.tasks import run_analysis_task

        missing_fields = []
        for field_name, field_label in HVAC_MANDATORY_FIELDS:
            if not (request.POST.get(field_name) or "").strip():
                missing_fields.append(field_label)

        if missing_fields:
            messages.error(
                request,
                "Please fill all mandatory fields before submitting: " + ", ".join(missing_fields),
            )
            return render(request, "procurement/request_create_hvac.html", _hvac_create_context())

        attrs_data = []
        attr_codes = request.POST.getlist("attr_code[]")
        attr_labels = request.POST.getlist("attr_label[]")
        attr_values = request.POST.getlist("attr_value[]")
        attr_types = request.POST.getlist("attr_type[]")
        for i, code in enumerate(attr_codes):
            if code.strip():
                attrs_data.append({
                    "attribute_code": code.strip(),
                    "attribute_label": attr_labels[i] if i < len(attr_labels) else code,
                    "data_type": attr_types[i] if i < len(attr_types) else "TEXT",
                    "value_text": attr_values[i] if i < len(attr_values) else "",
                    "is_required": False,
                })

        try:
            proc_request = ProcurementRequestService.create_request(
                title=request.POST.get("title", ""),
                description=request.POST.get("description", ""),
                domain_code="HVAC",                           # always HVAC
                schema_code="HVAC_GCC_V1",                    # HVAC schema
                request_type=request.POST.get(
                    "request_type", ProcurementRequestType.RECOMMENDATION
                ),
                priority=request.POST.get("priority", "HIGH"),
                geography_country=request.POST.get("geography_country", "UAE"),
                geography_city=request.POST.get("geography_city", ""),
                currency=request.POST.get("currency", "AED"),
                created_by=request.user,
                attributes=attrs_data if attrs_data else None,
            )

            store_id = (request.POST.get("f_store_id") or "").strip()
            if store_id:
                defaults = _build_hvac_store_profile_defaults(request.POST)
                defaults["updated_by"] = request.user
                profile, created = HVACStoreProfile.objects.update_or_create(
                    store_id=store_id,
                    defaults=defaults,
                )
                if created:
                    profile.created_by = request.user
                    profile.save(update_fields=["created_by"])

            run = AnalysisRunService.create_run(
                request=proc_request,
                run_type=AnalysisRunType.RECOMMENDATION,
                triggered_by=request.user,
            )
            run_analysis_task.delay(run.pk)

            # Generate market intelligence in parallel (best-effort, non-blocking)
            try:
                from apps.procurement.tasks import generate_market_intelligence_task
                generate_market_intelligence_task.apply_async(
                    args=[proc_request.pk],
                    countdown=20,  # give recommendation a 20s head start
                )
            except Exception as _mi_exc:
                logger.warning("Could not queue market intelligence task for pk=%s: %s", proc_request.pk, _mi_exc)

            messages.success(
                request,
                (
                    f"HVAC Procurement Request '{proc_request.title}' created successfully. "
                    "Recommendation analysis started."
                ),
            )
            return redirect("procurement:request_workspace", pk=proc_request.pk)
        except Exception as exc:
            logger.exception("hvac_create failed: %s", exc)
            messages.error(request, f"Failed to create HVAC request: {exc}")

    return render(request, "procurement/request_create_hvac.html", _hvac_create_context())


# ---------------------------------------------------------------------------
# HVAC Document Analyze Endpoint
# ---------------------------------------------------------------------------

def _map_hvac_extraction_to_fields(result: dict) -> dict:
    """Map AzureDIExtractorAgent output to HVAC form field names.

    Looks in result["header"] first, then falls back to result["key_value_pairs"]
    using fuzzy label matching.  Returns a flat dict of form_field -> value.
    """
    import re

    header = result.get("header") or {}
    kv_pairs = result.get("key_value_pairs") or []

    def _unwrap(v):
        """If the agent returned {'value': ..., 'confidence': ...}, extract the value string."""
        if isinstance(v, dict):
            v = v.get("value") or v.get("text") or v.get("content") or ""
        return str(v).strip() if v is not None else ""

    # Build a normalised lookup from kv_pairs: {lower_key: value}
    kv_lookup: dict = {}
    for kv in kv_pairs:
        raw_key = (kv.get("key") or kv.get("label") or "").lower().strip()
        raw_val = _unwrap(kv.get("value") or "")
        if raw_key and raw_val:
            kv_lookup[raw_key] = raw_val

    def _get(header_keys, kv_fragments, default=""):
        """Try header dict first, then kv_lookup, then KV fragment scanning."""
        for k in header_keys:
            v = header.get(k)
            if v is not None:
                v = _unwrap(v)
                if v:
                    return v
        norm_frags = [f.lower() for f in kv_fragments]
        for k, v in kv_lookup.items():
            if any(frag in k for frag in norm_frags):
                if v:
                    return v
        return default

    def _numeric(val, fallback=""):
        """Extract first number from a string."""
        if not val:
            return fallback
        m = re.search(r"[\d]+(?:\.\d+)?", str(val).replace(",", ""))
        return m.group(0) if m else fallback

    def _normalise_priority(val):
        v = val.upper() if val else ""
        if "CRITICAL" in v:
            return "CRITICAL"
        if "HIGH" in v:
            return "HIGH"
        if "LOW" in v:
            return "LOW"
        return "MEDIUM"

    def _normalise_store_type(val):
        v = val.upper() if val else ""
        if "MALL" in v:
            return "MALL"
        if "WAREHOUSE" in v or "WH" in v:
            return "WAREHOUSE"
        if "OFFICE" in v:
            return "OFFICE"
        if "STANDALONE" in v:
            return "STANDALONE"
        if "DATA" in v and "CENTER" in v:
            return "DATA_CENTER"
        if "RESTAURANT" in v or "F&B" in v or "FNB" in v or "FOOD" in v or "BEVERAGE" in v:
            return "RESTAURANT"
        return val if val in ("MALL", "STANDALONE", "WAREHOUSE", "OFFICE", "DATA_CENTER", "RESTAURANT") else ""

    def _normalise_store_format(val):
        v = val.upper() if val else ""
        if "HYPER" in v:
            return "HYPERMARKET"
        if "FURNITURE" in v:
            return "FURNITURE"
        if "ELECTRONICS" in v:
            return "ELECTRONICS"
        if "FOOD" in v or "BEVERAGE" in v or "F&B" in v:
            return "FOOD_BEVERAGE"
        if "RETAIL" in v:
            return "RETAIL"
        return ""

    def _normalise_level(val):
        v = val.upper() if val else ""
        if "HIGH" in v:
            return "HIGH"
        if "LOW" in v:
            return "LOW"
        if "MEDIUM" in v or "MED" in v:
            return "MEDIUM"
        return ""

    def _normalise_country(val):
        v = val.upper() if val else ""
        if "UAE" in v or "UNITED ARAB" in v or "EMIRATES" in v:
            return "UAE"
        if "KSA" in v or "SAUDI" in v or "KINGDOM" in v:
            return "KSA"
        if "QATAR" in v:
            return "QATAR"
        if "OMAN" in v:
            return "OMAN"
        if "KUWAIT" in v:
            return "KUWAIT"
        if "BAHRAIN" in v:
            return "BAHRAIN"
        return val or ""

    def _normalise_currency(val):
        v = val.upper() if val else ""
        if "SAR" in v:
            return "SAR"
        if "OMR" in v:
            return "OMR"
        if "QAR" in v:
            return "QAR"
        if "KWD" in v:
            return "KWD"
        if "BHD" in v:
            return "BHD"
        if "USD" in v:
            return "USD"
        return "AED"

    raw_title = _get(
        ["title", "request_title"],
        ["request title", "title"],
    )
    raw_priority = _get(["priority"], ["priority"])
    raw_type = _get(["request_type", "type"], ["request type", "type"])
    raw_desc = _get(["description", "background", "description_background"], ["description", "background"])
    raw_currency = _get(["currency"], ["currency"])
    raw_store_id = _get(["store_id", "facility_id", "store_number"], ["store id", "facility id", "store number", "facility identification"])
    raw_brand = _get(["brand", "brand_name"], ["brand"])
    raw_country = _get(["country", "geography_country"], ["country"])
    raw_city = _get(["city", "geography_city"], ["city"])
    raw_store_type = _get(["store_type", "facility_type"], ["store type", "facility type"])
    raw_store_format = _get(["store_format", "format"], ["store format", "format"])
    raw_area = _get(["area_sqft", "area", "area_sq_ft"], ["area (sq ft)", "area sq ft", "area sqft", "total area"])
    raw_ceiling = _get(["ceiling_height_ft", "ceiling_height", "ceiling"], ["ceiling height", "ceiling"])
    raw_hours = _get(["operating_hours", "hours"], ["operating hours"])
    raw_footfall = _get(["footfall_category", "footfall"], ["footfall category", "footfall"])
    raw_temp = _get(["ambient_temp_max", "ambient_temperature_max", "ambient_temp", "temperature_max"],
                    ["ambient temp", "ambient temperature", "temperature max"])
    raw_humidity = _get(["humidity_level", "humidity"], ["humidity level", "humidity"])
    raw_dust = _get(["dust_exposure", "dust"], ["dust exposure", "dust"])
    raw_heat = _get(["heat_load_category", "heat_load"], ["heat load category", "heat load"])
    raw_fresh_air = _get(["fresh_air_requirement", "fresh_air"], ["fresh air requirement", "fresh air"])
    raw_budget = _get(["budget_level", "budget"], ["budget level", "budget"])
    raw_energy = _get(
        ["energy_efficiency_priority", "energy_priority", "energy_efficiency"],
        ["energy efficiency priority", "energy priority", "energy efficiency"],
    )
    raw_landlord = _get(["landlord_constraint", "landlord_constraints"], ["landlord constraint"])

    fields = {}
    if raw_title:
        fields["title"] = raw_title
    if raw_priority:
        fields["priority"] = _normalise_priority(raw_priority)
    if raw_type:
        # Normalize to RECOMMENDATION by default for HVAC
        rt = raw_type.upper()
        fields["request_type"] = "RECOMMENDATION" if "RECOM" in rt else raw_type
    if raw_desc:
        fields["description"] = raw_desc
    if raw_currency:
        fields["currency"] = _normalise_currency(raw_currency)
    if raw_store_id:
        fields["f_store_id"] = raw_store_id
    if raw_brand:
        fields["f_brand"] = raw_brand
    if raw_country:
        fields["f_country"] = _normalise_country(raw_country)
    if raw_city:
        fields["f_city"] = raw_city
    if raw_store_type:
        fields["f_store_type"] = _normalise_store_type(raw_store_type)
    if raw_store_format:
        fields["f_store_format"] = _normalise_store_format(raw_store_format)
    if raw_area:
        num = _numeric(raw_area)
        if num:
            fields["f_area_sqft"] = num
    if raw_ceiling:
        num = _numeric(raw_ceiling)
        if num:
            fields["f_ceiling_height_ft"] = num
    if raw_hours:
        # Try to match to select options
        h = raw_hours.lower()
        if "24" in h:
            fields["f_operating_hours"] = "24 Hours"
        elif "8" in h and "6" in h:
            fields["f_operating_hours"] = "8 AM - 6 PM"
        elif "9" in h and "10" in h:
            fields["f_operating_hours"] = "9 AM - 10 PM"
        elif "9" in h and "11" in h:
            fields["f_operating_hours"] = "9 AM - 11 PM"
        elif "10" in h and "10" in h:
            fields["f_operating_hours"] = "10 AM - 10 PM"
        elif "10" in h and "11" in h:
            fields["f_operating_hours"] = "10 AM - 11 PM"
        elif "10" in h and "12" in h:
            fields["f_operating_hours"] = "10 AM - 12 AM"
        elif "6" in h and "10" in h:
            fields["f_operating_hours"] = "6 AM - 10 PM"
        else:
            fields["f_operating_hours"] = raw_hours
    if raw_footfall:
        fields["f_footfall_category"] = _normalise_level(raw_footfall)
    if raw_temp:
        num = _numeric(raw_temp)
        if num:
            fields["f_ambient_temp_max"] = num
    if raw_humidity:
        fields["f_humidity_level"] = _normalise_level(raw_humidity)
    if raw_dust:
        fields["f_dust_exposure"] = _normalise_level(raw_dust)
    if raw_heat:
        fields["f_heat_load_category"] = _normalise_level(raw_heat)
    if raw_fresh_air:
        fields["f_fresh_air_requirement"] = _normalise_level(raw_fresh_air)
    if raw_budget:
        fields["f_budget_level"] = _normalise_level(raw_budget)
    if raw_energy:
        fields["f_energy_efficiency_priority"] = _normalise_level(raw_energy)
    if raw_landlord:
        fields["f_landlord_constraint"] = raw_landlord

    return fields


@login_required
@require_http_methods(["POST"])
def hvac_analyze_document(request):
    """Accept an uploaded HVAC document, run Azure DI extraction, and return mapped form fields as JSON."""
    from django.http import JsonResponse
    import mimetypes

    uploaded_file = request.FILES.get("document")
    if not uploaded_file:
        return JsonResponse({"success": False, "error": "No document file provided."}, status=400)

    # Supported formats by Azure DI
    allowed_mime = {
        "application/pdf",
        "image/jpeg", "image/jpg", "image/png",
        "image/tiff", "image/bmp",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    mime_type, _ = mimetypes.guess_type(uploaded_file.name)
    if not mime_type:
        ext = uploaded_file.name.rsplit(".", 1)[-1].lower() if "." in uploaded_file.name else ""
        mime_map = {
            "pdf": "application/pdf",
            "jpg": "image/jpeg", "jpeg": "image/jpeg",
            "png": "image/png", "tiff": "image/tiff", "tif": "image/tiff",
            "bmp": "image/bmp",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "html": "text/html",
        }
        mime_type = mime_map.get(ext, "application/octet-stream")

    # We also allow HTML (for the sample forms we provide in media/hvac_forms/)
    allowed_mime.add("text/html")

    if mime_type not in allowed_mime:
        return JsonResponse({
            "success": False,
            "error": f"Unsupported file type '{mime_type}'. Please upload a PDF, image, DOCX, XLSX, or HTML file.",
        }, status=400)

    try:
        file_bytes = uploaded_file.read()
    except Exception as exc:
        return JsonResponse({"success": False, "error": f"Failed to read uploaded file: {exc}"}, status=500)

    # For HTML files (sample forms), pass them as text + instruct LLM to extract from HTML content
    if mime_type == "text/html":
        # Azure DI does not support HTML natively -- convert to plain text extraction via GPT only
        try:
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(file_bytes.decode("utf-8", errors="ignore"), "html.parser")
            plain_text = soup.get_text(separator="\n", strip=True)
        except Exception:
            plain_text = file_bytes.decode("utf-8", errors="ignore")

        # Build a direct GPT-based extraction bypassing Azure DI
        try:
            from apps.agents.services.llm_client import LLMClient, LLMMessage
            llm = LLMClient(max_tokens=4096)
            system_prompt = (
                "You are an HVAC procurement document parser. "
                "Extract all structured data from the following HVAC procurement request form text. "
                "Return a JSON object with these keys: "
                "title, priority, request_type, description, currency, store_id, brand, country, city, "
                "store_type, store_format, area_sqft, ceiling_height_ft, operating_hours, footfall_category, "
                "ambient_temp_max, humidity_level, dust_exposure, heat_load_category, fresh_air_requirement, "
                "landlord_constraint. "
                "Use null for fields not found. Return only valid JSON."
            )
            response = llm.chat(messages=[
                LLMMessage(role="system", content=system_prompt),
                LLMMessage(role="user", content=f"Document text:\n\n{plain_text[:12000]}"),
            ])
            import json as _json
            content = (response.content or "").strip()
            if content.startswith("```"):
                content = "\n".join(l for l in content.splitlines() if not l.strip().startswith("```")).strip()
            extracted_header = _json.loads(content)
            result = {
                "success": True,
                "doc_type": "hvac_request_form",
                "confidence": 0.82,
                "header": extracted_header,
                "line_items": [],
                "commercial_terms": {},
                "key_value_pairs": [],
                "engine": "gpt4o_html",
                "duration_ms": 0,
                "error": None,
            }
        except Exception as exc:
            logger.warning("HTML HVAC extraction failed: %s", exc)
            return JsonResponse({"success": False, "error": f"Document analysis failed: {exc}"}, status=500)
    else:
        # Full Azure DI + GPT extraction
        try:
            from apps.procurement.agents.Azure_Document_Intelligence_Extractor_Agent import AzureDIExtractorAgent
            result = AzureDIExtractorAgent.extract(
                file_bytes=file_bytes,
                mime_type=mime_type,
                doc_type_hint="hvac_request_form",
            )
        except Exception as exc:
            logger.warning("AzureDIExtractorAgent call failed: %s", exc)
            return JsonResponse({"success": False, "error": f"Document analysis failed: {exc}"}, status=500)

    if not result.get("success"):
        return JsonResponse({
            "success": False,
            "error": result.get("error") or "Extraction returned no data.",
            "confidence": result.get("confidence", 0.0),
        }, status=422)

    # Map extracted fields to HVAC form fields
    fields = _map_hvac_extraction_to_fields(result)

    return JsonResponse({
        "success": True,
        "fields": fields,
        "confidence": result.get("confidence", 0.0),
        "doc_type": result.get("doc_type", "unknown"),
        "engine": result.get("engine", ""),
        "duration_ms": result.get("duration_ms", 0),
        "fields_extracted": len(fields),
    })


@login_required
@permission_required_code("procurement.create")
def api_hvac_store_suggestions(request):
    """Return Store ID suggestions for HVAC form autosuggest/autofill."""
    query = (request.GET.get("q") or "").strip()

    qs = HVACStoreProfile.objects.filter(is_active=True)
    if query:
        qs = qs.filter(store_id__icontains=query)

    limit = 50 if query else 500
    qs = qs.order_by("store_id")[:limit]

    results = []
    for profile in qs:
        results.append({
            "store_id": profile.store_id,
            "brand": profile.brand,
            "country": profile.country,
            "city": profile.city,
            "store_type": profile.store_type,
            "store_format": profile.store_format,
            "area_sqft": profile.area_sqft,
            "ceiling_height_ft": profile.ceiling_height_ft,
            "operating_hours": profile.operating_hours,
            "footfall_category": profile.footfall_category,
            "ambient_temp_max": profile.ambient_temp_max,
            "humidity_level": profile.humidity_level,
            "dust_exposure": profile.dust_exposure,
            "heat_load_category": profile.heat_load_category,
            "fresh_air_requirement": profile.fresh_air_requirement,
            "landlord_constraints": profile.landlord_constraints,
            "existing_hvac_type": profile.existing_hvac_type,
            "budget_level": profile.budget_level,
            "energy_efficiency_priority": profile.energy_efficiency_priority,
        })

    return JsonResponse({"results": results})


@login_required
@permission_required_code("procurement.create")
def api_hvac_store_create(request):
    """AJAX POST -- create a new HVACStoreProfile and return the full profile JSON."""
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    data = request.POST
    store_id = (data.get("store_id") or "").strip()
    if not store_id:
        return JsonResponse({"error": "Store ID is required."}, status=400)

    if HVACStoreProfile.objects.filter(store_id=store_id).exists():
        return JsonResponse({"error": f"Store ID '{store_id}' already exists. Please choose a different ID."},
                            status=409)

    try:
        defaults = _build_hvac_store_profile_defaults(data)
        profile = HVACStoreProfile.objects.create(
            store_id=store_id,
            created_by=request.user,
            **defaults,
        )
        result = {
            "store_id": profile.store_id,
            "brand": profile.brand,
            "country": profile.country,
            "city": profile.city,
            "store_type": profile.store_type,
            "store_format": profile.store_format,
            "area_sqft": profile.area_sqft,
            "ceiling_height_ft": profile.ceiling_height_ft,
            "operating_hours": profile.operating_hours,
            "footfall_category": profile.footfall_category,
            "ambient_temp_max": profile.ambient_temp_max,
            "humidity_level": profile.humidity_level,
            "dust_exposure": profile.dust_exposure,
            "heat_load_category": profile.heat_load_category,
            "fresh_air_requirement": profile.fresh_air_requirement,
            "landlord_constraints": profile.landlord_constraints,
            "existing_hvac_type": profile.existing_hvac_type,
            "budget_level": profile.budget_level,
            "energy_efficiency_priority": profile.energy_efficiency_priority,
        }
        return JsonResponse({"ok": True, "profile": result}, status=201)
    except Exception as exc:
        logger.exception("api_hvac_store_create failed: %s", exc)
        return JsonResponse({"error": str(exc)}, status=500)


# ---------------------------------------------------------------------------
# 3. Request Workspace (Detail / Deep Dive)
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view")
def request_workspace(request, pk):
    """Full workspace view for a procurement request."""
    proc_request = get_object_or_404(
        ProcurementRequest.objects.select_related("created_by", "assigned_to"),
        pk=pk,
    )
    attributes = proc_request.attributes.all()
    quotations = proc_request.quotations.select_related("uploaded_document").all()
    runs = proc_request.analysis_runs.select_related("triggered_by").order_by("-created_at")

    # Latest recommendation
    recommendation = RecommendationResult.objects.filter(
        run__request=proc_request,
    ).select_related("run").order_by("-created_at").first()

    # Latest benchmark results
    benchmarks = BenchmarkResult.objects.filter(
        run__request=proc_request,
    ).select_related("run", "quotation").prefetch_related("lines").order_by("-created_at")

    # Latest compliance
    compliance = ComplianceResult.objects.filter(
        run__request=proc_request,
    ).select_related("run").order_by("-created_at").first()

    # Activity timeline from existing governance
    audit_events = AuditService.fetch_entity_history("ProcurementRequest", proc_request.pk)

    # Latest validation result
    validation_result = (
        ValidationResult.objects
        .filter(run__request=proc_request)
        .select_related("run")
        .prefetch_related("items")
        .order_by("-created_at")
        .first()
    )
    validation_items = validation_result.items.all() if validation_result else []

    # ReasonSummaryAgent -- structured explanation for the workspace UI.
    # Result is cached in recommendation.reason_summary_json to avoid hitting
    # the LLM API on every page load.  Cache is populated on first visit and
    # invalidated only when the user explicitly clicks "Regenerate".
    # Auto-invalidation: if cached conditions_table rows lack the "rule_filter"
    # key (added in the rule_conditions upgrade), the cache is stale and is
    # rebuilt transparently.
    reason_summary = None
    if recommendation:
        _cached = recommendation.reason_summary_json
        _cache_stale = False
        if _cached:
            _ct = _cached.get("conditions_table") or []
            if _ct and "rule_filter" not in (_ct[0] if _ct else {}):
                _cache_stale = True

        if _cached and not _cache_stale:
            # Serve from cache -- no LLM call needed
            reason_summary = _cached
        else:
            try:
                reason_summary = ReasonSummaryAgent.generate(recommendation)
                if reason_summary:
                    recommendation.reason_summary_json = reason_summary
                    recommendation.save(update_fields=["reason_summary_json"])
            except Exception as _e:
                logger.warning("ReasonSummaryAgent failed for request %s: %s", pk, _e)

    # Latest AI Market Intelligence suggestions (for Manual RFQ product picker)
    from apps.procurement.models import MarketIntelligenceSuggestion, HVACServiceScope
    mi_record = MarketIntelligenceSuggestion.objects.filter(
        request=proc_request
    ).order_by("-created_at").first()
    mi_suggestions = (mi_record.suggestions_json or []) if mi_record else []

    # Service scopes -- fetch all active rows; filter to match recommended system
    all_service_scopes = list(HVACServiceScope.objects.filter(is_active=True).order_by("sort_order", "system_type"))
    matched_service_scope = None

    # Keyword map: text patterns found in recommended_option or system_code -> HVACServiceScope.system_type
    _SCOPE_KEYWORD_MAP = [
        ("VRF",               "VRF"),
        ("VARIABLE REFRIGERANT", "VRF"),
        ("CHILLER",           "CHILLER"),
        ("CHILLED WATER",     "CHILLER"),
        ("FCU",               "FCU"),
        ("FAN COIL",          "FCU"),
        ("CASSETTE",          "CASSETTE"),
        ("PACKAGED_DX",       "PACKAGED_DX"),
        ("PACKAGED DX",       "PACKAGED_DX"),
        ("PACKAGED UNIT",     "PACKAGED_DX"),
        ("PACKAGED",          "PACKAGED_DX"),
        ("SPLIT_AC",          "SPLIT_AC"),
        ("SPLIT AC",          "SPLIT_AC"),
        ("SPLIT AIR",         "SPLIT_AC"),
        ("SPLIT",             "SPLIT_AC"),
    ]

    def _find_scope_by_code(code_upper, scopes):
        """Find a single HVACServiceScope by exact, startswith, or contains match."""
        for _ss in scopes:
            _st = _ss.system_type.upper()
            if _st == code_upper or code_upper.startswith(_st) or _st in code_upper:
                return _ss
        return None

    def _find_scope_by_text(text_upper, scopes):
        """Find a single HVACServiceScope by scanning text for known keywords."""
        for _kw, _sys in _SCOPE_KEYWORD_MAP:
            if _kw in text_upper:
                for _ss in scopes:
                    if _ss.system_type.upper() == _sys:
                        return _ss
        return None

    if recommendation:
        # Step 1: try reason_summary.system_code (startswith / contains match)
        if reason_summary:
            _sys_code = (getattr(reason_summary, "system_code", "") or "").upper()
            if _sys_code:
                matched_service_scope = _find_scope_by_code(_sys_code, all_service_scopes)

        # Step 2: if still no match, scan recommendation.recommended_option text
        if not matched_service_scope:
            _rec_text = (recommendation.recommended_option or "").upper()
            if _rec_text:
                matched_service_scope = _find_scope_by_text(_rec_text, all_service_scopes)

        # When a recommendation exists, do NOT show all scopes -- only the matched one
        # (if matched_service_scope is None the template will show a "not configured" notice)
        all_service_scopes = []

    # --- RFQ scope rows from HVACServiceScope (for RFQ modal Step 3 qty table) ---
    _all_scopes_for_rfq = HVACServiceScope.objects.filter(is_active=True).order_by("sort_order", "system_type")
    rfq_scope_json = {}
    for _ss in _all_scopes_for_rfq:
        _rows = []
        for _cat, _field_text in [
            ("Equipment",       _ss.equipment_scope),
            ("Installation",    _ss.installation_services),
            ("Piping/Ducting",  _ss.piping_ducting),
            ("Electrical",      _ss.electrical_works),
            ("Controls",        _ss.controls_accessories),
            ("Testing",         _ss.testing_commissioning),
        ]:
            for _line in (_field_text or "").splitlines():
                _line = _line.strip().lstrip("-*. ").strip()
                if _line:
                    _rows.append([_cat, _line, "LS", 1])
        if _rows:
            rfq_scope_json[_ss.system_type] = _rows

    # --- Products from DB for RFQ manual product selection (Step 2 cards) ---
    from apps.procurement.models import Product as _HVACProduct
    rfq_db_products = {}
    for _p in _HVACProduct.objects.filter(is_active=True).values(
        "system_type", "manufacturer", "product_name", "capacity_kw", "cop_rating", "sku"
    ).order_by("system_type", "manufacturer", "capacity_kw"):
        _stype = _p["system_type"]
        if _stype not in rfq_db_products:
            rfq_db_products[_stype] = []
        rfq_db_products[_stype].append({
            "code": _stype,
            "label": "{} {}".format(_p["manufacturer"], _p["product_name"]),
            "desc": "{}kW capacity".format(_p["capacity_kw"]),
            "brand": _p["manufacturer"],
            "capacity": str(_p["capacity_kw"]),
            "cop": str(_p["cop_rating"]) if _p["cop_rating"] else "",
            "sku": _p["sku"],
        })

    # --- Latest GeneratedRFQ for this request (for persistent download panel) ---
    from apps.procurement.models import GeneratedRFQ as _GeneratedRFQ
    rfq_record = (
        _GeneratedRFQ.objects
        .filter(request=proc_request)
        .order_by("-created_at")
        .first()
    )

    # Generate 2-hour SAS URLs for in-browser preview (fail silent)
    rfq_xlsx_view_url = ""
    rfq_pdf_view_url = ""
    if rfq_record:
        try:
            from apps.documents.blob_service import (
                generate_blob_sas_url as _gen_sas,
                is_blob_storage_enabled as _blob_ok,
            )
            if _blob_ok():
                if rfq_record.xlsx_blob_path:
                    rfq_xlsx_view_url = _gen_sas(rfq_record.xlsx_blob_path, expiry_minutes=120) or ""
                if rfq_record.pdf_blob_path:
                    rfq_pdf_view_url = _gen_sas(rfq_record.pdf_blob_path, expiry_minutes=120) or ""
        except Exception:
            pass

    return render(request, "procurement/request_workspace.html", {
        "proc_request": proc_request,
        "attributes": attributes,
        "quotations": quotations,
        "runs": runs,
        "recommendation": recommendation,
        "reason_summary": reason_summary,
        "benchmarks": benchmarks,
        "compliance": compliance,
        "validation_result": validation_result,
        "validation_items": validation_items,
        "audit_events": audit_events[:50],
        "status_choices": ProcurementRequestStatus.choices,
        "mi_suggestions": mi_suggestions,
        "mi_record": mi_record,
        "all_service_scopes": all_service_scopes,
        "matched_service_scope": matched_service_scope,
        "rfq_scope_json": rfq_scope_json,
        "rfq_db_products": rfq_db_products,
        "rfq_record": rfq_record,
        "rfq_xlsx_view_url": rfq_xlsx_view_url,
        "rfq_pdf_view_url": rfq_pdf_view_url,
    })


# ---------------------------------------------------------------------------
# 3b. Generate RFQ (Excel or PDF download)
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# 3c. Regenerate Reasoning Summary (clears cache and re-calls LLM)
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view")
def regenerate_reason_summary(request, pk):
    """POST-only: clear the cached reason_summary_json and regenerate via LLM."""
    if request.method != "POST":
        return redirect("procurement:request_workspace", pk=pk)

    proc_request = get_object_or_404(ProcurementRequest, pk=pk)
    recommendation = (
        RecommendationResult.objects
        .filter(run__request=proc_request)
        .order_by("-created_at")
        .first()
    )
    if recommendation:
        # Clear the cache so the next workspace load will re-call the LLM
        recommendation.reason_summary_json = None
        recommendation.save(update_fields=["reason_summary_json"])
        try:
            new_summary = ReasonSummaryAgent.generate(recommendation)
            if new_summary:
                recommendation.reason_summary_json = new_summary
                recommendation.save(update_fields=["reason_summary_json"])
            messages.success(request, "Reasoning summary regenerated successfully.")
        except Exception as exc:
            logger.warning("regenerate_reason_summary failed for pk=%s: %s", pk, exc)
            messages.error(request, "Regeneration failed -- please try again shortly.")
    else:
        messages.warning(request, "No recommendation found to regenerate.")

    return redirect("procurement:request_workspace", pk=pk)


@login_required
@permission_required_code("procurement.view")
def generate_rfq(request, pk):
    """Generate and save RFQ documents (Excel + PDF) to Azure Blob Storage.

    POST  (action=generate)  -- build both files, upload to blob, save
          GeneratedRFQ record, return JSON {"status":"ok","rfq_id":...}.

    GET   action=download&rfq_id=<id>&format=xlsx|pdf
          -- redirect to Azure Blob SAS URL for existing file.

    GET   (legacy, no action) -- build xlsx, stream as download.
    """
    import io
    import json as _json
    import datetime
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse("openpyxl is not installed.", status=500)

    from django.http import HttpResponse, JsonResponse

    # ------------------------------------------------------------------
    # action=download: serve existing file from blob (or 404)
    # ------------------------------------------------------------------
    if request.method == "GET" and request.GET.get("action") == "download":
        from apps.procurement.models import GeneratedRFQ as _GRFQ
        from apps.documents.blob_service import generate_blob_sas_url, is_blob_storage_enabled
        rfq_id = request.GET.get("rfq_id", "")
        dl_fmt = request.GET.get("format", "xlsx").strip().lower()
        try:
            _grfq = _GRFQ.objects.get(pk=int(rfq_id), request__pk=pk)
        except Exception:
            from django.http import Http404
            raise Http404("RFQ record not found.")
        blob_path = _grfq.xlsx_blob_path if dl_fmt == "xlsx" else _grfq.pdf_blob_path
        if blob_path and is_blob_storage_enabled():
            # Force download (not inline preview) by setting Content-Disposition via SAS rscd
            _rfq_ref = f"RFQ-{pk:04d}"
            if dl_fmt == "pdf":
                _cd = f'attachment; filename="{_rfq_ref}.pdf"'
            else:
                _cd = f'attachment; filename="{_rfq_ref}.xlsx"'
            sas_url = generate_blob_sas_url(blob_path, expiry_minutes=30, content_disposition=_cd)
            if sas_url:
                from django.shortcuts import redirect as _redirect
                return _redirect(sas_url)
        # Blob path unavailable -- fall through to regenerate + stream
        _dl_fmt_override = dl_fmt
    else:
        _dl_fmt_override = None  # not a download-only GET

    # ------------------------------------------------------------------
    # Parse params (POST for generate action, GET for legacy)
    # ------------------------------------------------------------------
    _is_generate_post = (request.method == "POST")
    if _is_generate_post:
        product_param_raw = request.POST.get("product", "RECOMMENDED")
        raw_qty = request.POST.get("qty_json", "")
    else:
        product_param_raw = request.GET.get("product", "RECOMMENDED")
        raw_qty = request.GET.get("qty_json", "")

    # qty overrides: JSON dict {"row_index": qty}
    qty_overrides = {}
    try:
        if raw_qty:
            qty_overrides = {int(k): v for k, v in _json.loads(raw_qty).items()}
    except Exception:
        qty_overrides = {}

    product_param = product_param_raw.strip().upper()

    proc_request = get_object_or_404(
        ProcurementRequest.objects.select_related("created_by"),
        pk=pk,
    )

    # -----------------------------------------------------------------------
    # Delegate all generation logic to RFQGeneratorAgent.
    #
    # When the user clicks "Use Recommendation" in the modal form the view
    # receives product_param == "RECOMMENDED" and passes selection_mode=
    # "RECOMMENDED" to the agent, which fetches the latest RecommendationResult
    # and uses its system_type_code automatically.
    # -----------------------------------------------------------------------
    from apps.procurement.agents.RFQ_Generator_Agent import RFQGeneratorAgent

    rfq_result = RFQGeneratorAgent.run(
        proc_request,
        selection_mode=product_param,        # "RECOMMENDED" or a system code
        qty_overrides=qty_overrides,
        generated_by=request.user,
        save_record=_is_generate_post,       # only persist on POST action=generate
    )

    if rfq_result.error and not rfq_result.xlsx_bytes:
        # Hard failure: nothing was built
        from django.http import HttpResponseServerError
        logger.error("RFQGeneratorAgent hard failure for pk=%s: %s", pk, rfq_result.error)
        if _is_generate_post:
            return JsonResponse({"status": "error", "detail": rfq_result.error}, status=500)
        return HttpResponseServerError("RFQ generation failed -- please try again.")

    xlsx_bytes   = rfq_result.xlsx_bytes
    pdf_bytes    = rfq_result.pdf_bytes
    rfq_ref      = rfq_result.rfq_ref
    filename_xlsx = rfq_result.filename_xlsx
    filename_pdf  = rfq_result.filename_pdf

    # ---- Short-circuit: all generation handled by the agent above. ----
    # Return immediately so legacy inline code below is never reached.

    if _is_generate_post:
        _rfq_rec = rfq_result.rfq_record
        return JsonResponse({
            "status": "ok",
            "rfq_ref": rfq_ref,
            "rfq_id": _rfq_rec.pk if _rfq_rec else None,
            "has_xlsx": bool(xlsx_bytes),
            "has_pdf": bool(pdf_bytes),
            "blob_enabled": bool(
                _rfq_rec and _rfq_rec.xlsx_blob_path
            ) if _rfq_rec else False,
        })

    if _dl_fmt_override == "pdf":
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename_pdf}"'
        return resp

    response = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename_xlsx}"'
    return response

    # =========================================================================
    # NOTE: Code below this point is unreachable (kept for reference only).
    # All generation logic now lives in RFQ_Generator_Agent.py.
    # =========================================================================
    SYSTEM_SCOPE = {
        "VRF": {
            "capacity": "As per heat load calculation (TR)",
            "scope": [
                ("Equipment",    "VRF Outdoor Unit(s)",                             "Nos",  ""),
                ("Equipment",    "VRF Indoor Units (Ceiling / Wall type per layout)","Nos",  ""),
                ("Piping",       "Refrigerant Copper Piping & Fittings",             "RM",   ""),
                ("Piping",       "Thermal Insulation for Refrigerant Pipes",         "RM",   ""),
                ("Electrical",   "Power Cabling & Distribution Boards",              "LS",   1),
                ("Controls",     "Central Controller / BMS Integration",             "LS",   1),
                ("Civil & MEP",  "Structural Support, Drainage, Penetrations",       "LS",   1),
                ("Installation", "Complete Installation Works",                      "LS",   1),
                ("Testing",      "Testing, Commissioning & Handover",                "LS",   1),
            ],
        },
        "SPLIT_AC": {
            "capacity": "As per heat load calculation (TR)",
            "scope": [
                ("Equipment",    "Split AC Outdoor Condensing Units",                "Nos",  ""),
                ("Equipment",    "Split AC Indoor Units",                            "Nos",  ""),
                ("Piping",       "Refrigerant Copper Piping & Fittings",             "RM",   ""),
                ("Piping",       "Condensate Drain Piping",                          "RM",   ""),
                ("Electrical",   "Power Cabling & MCB Distribution",                 "LS",   1),
                ("Civil & MEP",  "Structural Support & Wall Penetrations",           "LS",   1),
                ("Installation", "Complete Installation Works",                      "LS",   1),
                ("Testing",      "Testing, Commissioning & Handover",                "LS",   1),
            ],
        },
        "PACKAGED_DX": {
            "capacity": "As per heat load calculation (TR)",
            "scope": [
                ("Equipment",    "Packaged DX Unit(s) (Roof-Top / Split Packaged)", "Nos",  ""),
                ("Ducting",      "GI Ducting (Supply & Return)",                    "Sqm",  ""),
                ("Ducting",      "Flexible Duct Connectors",                        "RM",   ""),
                ("Diffusers",    "Supply / Return Air Diffusers & Grilles",         "Nos",  ""),
                ("Insulation",   "Duct Thermal & Acoustic Insulation",              "Sqm",  ""),
                ("Electrical",   "Power Cabling & Distribution Boards",             "LS",   1),
                ("Civil & MEP",  "Roof Curb, Support Structure, Penetrations",      "LS",   1),
                ("Installation", "Complete Installation Works",                     "LS",   1),
                ("Testing",      "Testing, Commissioning & Handover",               "LS",   1),
            ],
        },
        "CHILLER": {
            "capacity": "As per heat load calculation (TR)",
            "scope": [
                ("Equipment",    "Water-Cooled / Air-Cooled Chiller Plant",         "Nos",  ""),
                ("Equipment",    "Cooling Towers (if water-cooled)",                "Nos",  ""),
                ("Equipment",    "Air Handling Units (AHUs) / Fan Coil Units",      "Nos",  ""),
                ("Piping",       "Chilled Water & Condenser Water Piping",          "RM",   ""),
                ("Piping",       "Thermal Insulation for Chilled Water Pipes",      "RM",   ""),
                ("Pumps",        "Primary & Secondary Chilled Water Pumps",         "Nos",  ""),
                ("Electrical",   "LV Panels, Cabling & MCC",                        "LS",   1),
                ("Controls",     "BMS / DDC Control System",                        "LS",   1),
                ("Installation", "Complete Installation Works",                     "LS",   1),
                ("Testing",      "Testing, Commissioning & Handover",               "LS",   1),
            ],
        },
        "FCU": {
            "capacity": "As per heat load calculation (TR)",
            "scope": [
                ("Equipment",    "Fan Coil Units (2-pipe or 4-pipe)",               "Nos",  ""),
                ("Piping",       "Chilled Water Supply & Return Piping",            "RM",   ""),
                ("Piping",       "Condensate Drain Piping",                         "RM",   ""),
                ("Insulation",   "Pipe Thermal Insulation",                         "RM",   ""),
                ("Electrical",   "Power Cabling & Wiring",                          "LS",   1),
                ("Controls",     "Thermostat & Zone Controls",                      "LS",   1),
                ("Civil & MEP",  "Ceiling Works & Support Structures",              "LS",   1),
                ("Installation", "Complete Installation Works",                     "LS",   1),
                ("Testing",      "Testing, Commissioning & Handover",               "LS",   1),
            ],
        },
        "CASSETTE": {
            "capacity": "As per heat load calculation (TR)",
            "scope": [
                ("Equipment",    "Cassette Type Indoor Units (4-way blow)",         "Nos",  ""),
                ("Equipment",    "Outdoor Condensing Units",                        "Nos",  ""),
                ("Piping",       "Refrigerant Copper Piping & Fittings",            "RM",   ""),
                ("Piping",       "Condensate Drain Piping",                         "RM",   ""),
                ("Electrical",   "Power Cabling & Distribution",                    "LS",   1),
                ("Civil & MEP",  "Ceiling Cutouts, Diffuser Frames, Supports",      "LS",   1),
                ("Installation", "Complete Installation Works",                     "LS",   1),
                ("Testing",      "Testing, Commissioning & Handover",               "LS",   1),
            ],
        },
    }

    # ---- Normalize raw system text to a canonical DB system_type key ----
    _SCOPE_CODE_MAP = [
        ("VRF",               "VRF"),
        ("VARIABLE REFRIGERANT", "VRF"),
        ("CHILLER",           "CHILLER"),
        ("CHILLED WATER",     "CHILLER"),
        ("FCU",               "FCU"),
        ("FAN COIL",          "FCU"),
        ("CASSETTE",          "CASSETTE"),
        ("PACKAGED_DX",       "PACKAGED_DX"),
        ("PACKAGED DX",       "PACKAGED_DX"),
        ("PACKAGED UNIT",     "PACKAGED_DX"),
        ("PACKAGED",          "PACKAGED_DX"),
        ("SPLIT_AC",          "SPLIT_AC"),
        ("SPLIT AC",          "SPLIT_AC"),
        ("SPLIT AIR",         "SPLIT_AC"),
        ("SPLIT",             "SPLIT_AC"),
    ]

    def _normalize_system_code(raw):
        """Resolve any text form of an HVAC system name to its canonical DB system_type key."""
        if not raw:
            return "PACKAGED_DX"
        _u = raw.strip().upper()
        # 1. Direct match
        for _kw, _code in _SCOPE_CODE_MAP:
            if _u == _kw:
                return _code
        # 2. Startswith
        for _kw, _code in _SCOPE_CODE_MAP:
            if _u.startswith(_kw):
                return _code
        # 3. Contains
        for _kw, _code in _SCOPE_CODE_MAP:
            if _kw in _u:
                return _code
        return _u  # Return as-is -- may be a valid DB key already

    # ---- Resolve product ----
    # NOTE: use product_param that was already resolved from POST or GET above;
    # do NOT re-read from request.GET here as that would ignore POST values.
    if product_param == "RECOMMENDED":
        rec = RecommendationResult.objects.filter(
            run__request=proc_request,
        ).order_by("-created_at").first()
        if rec:
            _raw_sys = (
                (rec.output_payload_json or {}).get("system_type_code", "")
                or str(rec.recommended_option or "").split("(")[0].strip()
            )
            system_code = _normalize_system_code(_raw_sys)
            rationale = rec.reasoning_summary or "Based on store profile and site conditions."
            confidence_pct = round(float(rec.confidence_score or 0) * 100)
        else:
            system_code = "PACKAGED_DX"
            rationale = "Default recommendation -- no analysis run yet."
            confidence_pct = 0
        selection_basis = "AI / Rules Engine Recommendation"
    else:
        system_code = _normalize_system_code(product_param)
        rationale = "Manually selected based on project requirements."
        confidence_pct = 0
        selection_basis = "Manual Selection"

    # -- Scope: DB-first (HVACServiceScope mandatory), fallback to hardcoded SYSTEM_SCOPE --
    # DB scope takes absolute priority -- all 6 categories are always emitted
    # so that Equipment/Installation/Piping/Electrical/Controls/Testing are
    # always present in both the Excel and PDF documents.
    from apps.procurement.models import HVACServiceScope as _HSSModel
    _db_scope = _HSSModel.objects.filter(system_type=system_code, is_active=True).first()
    if not _db_scope:
        # Try a case-insensitive search as a safety net
        _db_scope = _HSSModel.objects.filter(
            system_type__iexact=system_code, is_active=True
        ).first()
    if _db_scope:
        system_label = _db_scope.display_name or SYSTEM_LABELS.get(system_code, system_code)
        _raw_rows = []
        for _cat, _field_text in [
            ("Equipment",       _db_scope.equipment_scope),
            ("Installation",    _db_scope.installation_services),
            ("Piping/Ducting",  _db_scope.piping_ducting),
            ("Electrical",      _db_scope.electrical_works),
            ("Controls",        _db_scope.controls_accessories),
            ("Testing",         _db_scope.testing_commissioning),
        ]:
            _lines_added = 0
            for _line in (_field_text or "").splitlines():
                _line = _line.strip().lstrip("-*. ").strip()
                if _line:
                    _raw_rows.append((_cat, _line, "LS", 1))
                    _lines_added += 1
            # Always emit at least one placeholder row per category so every
            # category header appears in the document even if the DB field is blank.
            if _lines_added == 0:
                _raw_rows.append((_cat, "(As per site conditions)", "LS", 1))
        capacity_note = "As per heat load calculation (TR)"
        scope_rows = _raw_rows
        logger.info(
            "RFQ scope for pk=%s system=%s: loaded %d rows from HVACServiceScope DB",
            pk, system_code, len(scope_rows),
        )
    else:
        system_label = SYSTEM_LABELS.get(system_code, system_code)
        scope_data = SYSTEM_SCOPE.get(system_code, SYSTEM_SCOPE["PACKAGED_DX"])
        capacity_note = scope_data["capacity"]
        scope_rows = list(scope_data["scope"])
        logger.info(
            "RFQ scope for pk=%s system=%s: DB scope not found, using hardcoded table",
            pk, system_code,
        )
    # Apply qty overrides
    if qty_overrides:
        scope_rows = [
            (cat, desc, unit, qty_overrides.get(i, qty))
            for i, (cat, desc, unit, qty) in enumerate(scope_rows)
        ]

    # ---- Gather attributes ----
    attr_map = {a.attribute_code: a.value_text for a in proc_request.attributes.all()}
    country       = attr_map.get("country", "") or proc_request.geography_country or ""
    city          = attr_map.get("city", "") or proc_request.geography_city or ""
    store_type    = attr_map.get("store_type", "")
    area_sqft     = attr_map.get("area_sqft", "")
    ceiling_h     = attr_map.get("ceiling_height", "")
    ambient       = attr_map.get("ambient_temp_max", "")
    humidity      = attr_map.get("humidity_level", "")
    cooling_tr    = attr_map.get("estimated_cooling_tr", "")
    budget        = attr_map.get("budget_level", "")

    today     = datetime.date.today().strftime("%d-%b-%Y")
    rfq_ref   = f"RFQ-{proc_request.pk:04d}-{datetime.date.today().strftime('%Y%m%d')}"
    capacity_display = f"{cooling_tr} TR" if cooling_tr else capacity_note

    # ================================================================ Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RFQ"
    ws.sheet_view.showGridLines = False  # clean white canvas

    # Column widths: A(idx)=6, B(param)=28, C(value/desc)=48, D(unit)=14, E(qty)=10
    for i, w in zip(range(1, 6), [6, 28, 48, 14, 10]):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Style helpers ----
    def _thin():
        s = Side(style="thin", color="BDBDBD")
        return Border(left=s, right=s, top=s, bottom=s)

    def _fill(hex_val):
        return PatternFill("solid", fgColor=hex_val)

    NAVY   = "1A3C5E"
    LTBLUE = "EAF2FB"
    GREY   = "F5F5F5"
    WHITE  = "FFFFFF"
    SECBG  = "D6E4F0"

    def _merge_write(row, col_start, col_end, value, font=None, fill=None,
                     align=None, height=None, border=True):
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=value)
        if font:  c.font = font
        if fill:  c.fill = fill
        if align: c.alignment = align
        if border:
            for col in range(col_start, col_end + 1):
                ws.cell(row=row, column=col).border = _thin()
        if height:
            ws.row_dimensions[row].height = height

    def _kv(row, label, value, label_fill=GREY):
        """Write a 2-column key-value row (columns B & C), blank A D E."""
        lc = ws.cell(row=row, column=2, value=label)
        lc.font = Font(bold=True, size=10)
        lc.fill = _fill(label_fill)
        lc.border = _thin()
        lc.alignment = Alignment(vertical="center")
        vc = ws.cell(row=row, column=3, value=value or "--")
        vc.font = Font(size=10)
        vc.fill = _fill(WHITE)
        vc.border = _thin()
        vc.alignment = Alignment(vertical="center", wrap_text=True)
        # Span value across C-E for readability
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        for col in range(3, 6):
            ws.cell(row=row, column=col).border = _thin()
        ws.row_dimensions[row].height = 18

    def _section_header(row, number, title):
        """Bold section title spanning B-E."""
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=5)
        c = ws.cell(row=row, column=2, value=f"{number}. {title}")
        c.font = Font(bold=True, size=11, color=NAVY)
        c.fill = _fill(SECBG)
        c.alignment = Alignment(vertical="center")
        c.border = _thin()
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = _thin()
        ws.row_dimensions[row].height = 20

    def _table_header(row, cols):
        """Write table column headers spanning B-E."""
        # cols = list of (col_index, label) with col_index relative to sheet
        for ci, label in cols:
            c = ws.cell(row=row, column=ci, value=label)
            c.font = Font(bold=True, size=10, color=NAVY)
            c.fill = _fill(LTBLUE)
            c.border = _thin()
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        ws.row_dimensions[row].height = 18

    # ============================== ROW 1: TITLE ==============================
    _merge_write(1, 1, 5,
        "REQUEST FOR QUOTATION (RFQ) - HVAC WORKS",
        font=Font(bold=True, size=16, color="FFFFFF"),
        fill=_fill(NAVY),
        align=Alignment(horizontal="center", vertical="center"),
        height=36, border=False)

    # Row 2: Ref + Date on same line
    ws.cell(row=2, column=2, value=f"RFQ Ref: {rfq_ref}").font = Font(bold=True, size=9, color="555555")
    ws.merge_cells("B2:C2")
    date_c = ws.cell(row=2, column=4, value=f"Date: {today}")
    date_c.font = Font(size=9, color="555555")
    date_c.alignment = Alignment(horizontal="right")
    ws.merge_cells("D2:E2")
    ws.row_dimensions[2].height = 14

    # ========================== INTRO PARAGRAPH ================================
    intro = (
        f"We invite your quotation for the Supply, Installation, Testing, and Commissioning (SITC) "
        f"of a {system_label} HVAC system for the store described below. "
        f"Please submit a detailed, itemised quotation covering all scope items listed in Section 3."
    )
    _merge_write(4, 2, 5, intro,
        font=Font(size=10),
        fill=_fill(WHITE),
        align=Alignment(wrap_text=True, vertical="top"),
        height=40, border=False)
    ws.row_dimensions[4].height = 44

    # ========================== SECTION 1: STORE DETAILS ======================
    r = 6
    _section_header(r, 1, "Store Details"); r += 1
    _table_header(r, [(2, "Parameter"), (3, "Value")]); r += 1
    store_rows = [
        ("Country",          country),
        ("City",             city),
        ("Store Type",       store_type),
        ("Area",             f"{area_sqft} sq ft" if area_sqft else ""),
        ("Ceiling Height",   f"{ceiling_h} ft" if ceiling_h else ""),
        ("Max Temperature",  f"{ambient} deg C" if ambient else ""),
        ("Humidity",         humidity),
        ("Budget Level",     budget),
    ]
    for param, val in store_rows:
        if not val:
            continue
        _kv(r, param, val); r += 1

    # ========================= SECTION 2: HVAC SYSTEM =========================
    r += 1
    _section_header(r, 2, "Recommended HVAC System"); r += 1
    _table_header(r, [(2, "Field"), (3, "Value")]); r += 1
    _kv(r, "System Type",        system_label); r += 1
    _kv(r, "Capacity",           capacity_display); r += 1
    _kv(r, "Selection Basis",    selection_basis); r += 1
    if confidence_pct:
        _kv(r, "Confidence", f"{confidence_pct}%"); r += 1
    # Rationale - taller row
    lc2 = ws.cell(row=r, column=2, value="Reason / Rationale")
    lc2.font = Font(bold=True, size=10)
    lc2.fill = _fill(GREY)
    lc2.border = _thin()
    lc2.alignment = Alignment(vertical="top")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    vc2 = ws.cell(row=r, column=3, value=rationale)
    vc2.font = Font(size=10, italic=True)
    vc2.fill = _fill(WHITE)
    vc2.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(2, 6):
        ws.cell(row=r, column=col).border = _thin()
    ws.row_dimensions[r].height = 50
    r += 1

    # ========================== SECTION 3: SCOPE OF WORK ======================
    r += 1
    _section_header(r, 3, "Scope of Work"); r += 1

    # Table header: A=S.No, B=Category, C=Description, D=Unit, E=Qty
    scope_hdr_cols = [(1, "S.No"), (2, "Category"), (3, "Description"), (4, "Unit"), (5, "Qty")]
    for ci, label in scope_hdr_cols:
        c = ws.cell(row=r, column=ci, value=label)
        c.font = Font(bold=True, size=10, color=NAVY)
        c.fill = _fill(LTBLUE)
        c.border = _thin()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    for sno, (cat, desc, unit, qty) in enumerate(scope_rows, 1):
        # S.No
        sno_c = ws.cell(row=r, column=1, value=sno)
        sno_c.font = Font(size=10)
        sno_c.fill = _fill(GREY)
        sno_c.border = _thin()
        sno_c.alignment = Alignment(horizontal="center", vertical="center")
        # Category
        cat_c = ws.cell(row=r, column=2, value=cat)
        cat_c.font = Font(bold=True, size=10)
        cat_c.fill = _fill(GREY)
        cat_c.border = _thin()
        cat_c.alignment = Alignment(vertical="center")
        # Description
        desc_c = ws.cell(row=r, column=3, value=desc)
        desc_c.font = Font(size=10)
        desc_c.fill = _fill(WHITE)
        desc_c.border = _thin()
        desc_c.alignment = Alignment(vertical="center", wrap_text=True)
        # Unit
        unit_c = ws.cell(row=r, column=4, value=unit)
        unit_c.font = Font(size=10)
        unit_c.fill = _fill(WHITE)
        unit_c.border = _thin()
        unit_c.alignment = Alignment(horizontal="center", vertical="center")
        # Qty
        qty_c = ws.cell(row=r, column=5, value=qty)
        qty_c.font = Font(size=10)
        qty_c.fill = _fill(WHITE)
        qty_c.border = _thin()
        qty_c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

    # ========================== SECTION 4: COMMERCIAL TERMS ===================
    r += 1
    _section_header(r, 4, "Commercial Terms"); r += 1
    _table_header(r, [(2, "Term"), (3, "Details")]); r += 1
    commercial_rows = [
        ("Quotation Validity", "90 days from submission date"),
        ("Submission Deadline", "As advised by issuing party"),
        ("Delivery / Completion", "Vendor to specify in quotation"),
        ("Payment Terms", "As per standard purchase order terms"),
        ("Currency", f"{country} local currency or as agreed"),
        ("Pricing Basis", "Lump sum (supply + installation + commissioning)"),
        ("Warranty", "Minimum 2 years on equipment and workmanship"),
        ("Compliance", "ASHRAE 90.1, SASO, and applicable local building code"),
    ]
    for term, detail in commercial_rows:
        _kv(r, term, detail); r += 1

    # ========================== SECTION 5: VENDOR PRICING RESPONSE ============
    r += 1
    _section_header(r, 5, "Vendor Pricing Response (To be filled by Vendor)"); r += 1
    vnd_cols = [(1, "S.No"), (2, "Category"), (3, "Description / Proposed Model"),
                (4, "Unit Price"), (5, "Total")]
    for ci, label in vnd_cols:
        c = ws.cell(row=r, column=ci, value=label)
        c.font = Font(bold=True, size=10, color=NAVY)
        c.fill = _fill(LTBLUE)
        c.border = _thin()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 18
    r += 1
    for idx in range(1, len(scope_rows) + 1):
        for ci in range(1, 6):
            c = ws.cell(row=r, column=ci, value=idx if ci == 1 else "")
            c.border = _thin()
            c.fill = _fill(WHITE)
            c.font = Font(size=10)
            if ci in (4, 5):
                c.alignment = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 18
        r += 1

    # Totals row
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    tc = ws.cell(row=r, column=3, value="GRAND TOTAL (Excl. VAT)")
    tc.font = Font(bold=True, size=10)
    tc.fill = _fill(SECBG)
    tc.border = _thin()
    tc.alignment = Alignment(horizontal="right", vertical="center")
    gt = ws.cell(row=r, column=5, value="")
    gt.border = _thin()
    gt.fill = _fill(SECBG)
    gt.font = Font(bold=True, size=10)
    ws.row_dimensions[r].height = 18
    r += 2

    # ========================== DISCLAIMER ====================================
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    dc = ws.cell(row=r, column=1,
        value=("Note: This RFQ is system-generated. All specifications are indicative."
               " A qualified HVAC engineer must review and confirm final scope before award."))
    dc.font = Font(size=8, italic=True, color="888888")
    dc.alignment = Alignment(wrap_text=True, horizontal="center")
    dc.fill = _fill("FFFDE7")
    ws.row_dimensions[r].height = 24

    # ================================================================ Response
    safe_title = "".join(
        c for c in (proc_request.title or "Request") if c.isalnum() or c in " _-"
    )[:30].strip().replace(" ", "_")

    # ---- Always build Excel buffer ----
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()
    filename_xlsx = f"RFQ_{rfq_ref}_{safe_title}.xlsx"

    # ---- Always build PDF buffer ----
    pdf_bytes = b""
    filename_pdf = f"RFQ_{rfq_ref}_{safe_title}.pdf"
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        )

        pdf_buf = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buf, pagesize=A4,
            leftMargin=18 * mm, rightMargin=18 * mm,
            topMargin=18 * mm, bottomMargin=18 * mm,
        )

        styles = getSampleStyleSheet()
        NAVY_RL  = colors.HexColor("#1A3C5E")
        LTBLUE_RL = colors.HexColor("#EAF2FB")
        SECBG_RL  = colors.HexColor("#D6E4F0")
        GREY_RL   = colors.HexColor("#F5F5F5")

        title_style = ParagraphStyle(
            "rfqTitle", parent=styles["Title"],
            fontSize=14, textColor=colors.white,
            backColor=NAVY_RL, spaceAfter=4, spaceBefore=0,
            alignment=1,  # center
        )
        sec_style = ParagraphStyle(
            "rfqSec", parent=styles["Normal"],
            fontSize=11, textColor=NAVY_RL, fontName="Helvetica-Bold",
            spaceAfter=2, spaceBefore=6,
        )
        normal_sm = ParagraphStyle(
            "rfqNorm", parent=styles["Normal"],
            fontSize=9, spaceAfter=1,
        )
        italic_sm = ParagraphStyle(
            "rfqItalic", parent=styles["Normal"],
            fontSize=9, fontName="Helvetica-Oblique", spaceAfter=4,
        )

        story = []

        # Title
        story.append(Paragraph("REQUEST FOR QUOTATION (RFQ) -- HVAC WORKS", title_style))
        story.append(Spacer(1, 4 * mm))

        # Ref + Date
        story.append(Paragraph(
            f"<b>RFQ Ref:</b> {rfq_ref} &nbsp;&nbsp;&nbsp; <b>Date:</b> {today}",
            normal_sm,
        ))
        story.append(Spacer(1, 3 * mm))

        # Intro
        story.append(Paragraph(
            f"We invite your quotation for the Supply, Installation, Testing, and Commissioning "
            f"(SITC) of a <b>{system_label}</b> HVAC system for the store described below. "
            f"Please submit a detailed, itemised quotation covering all scope items in Section 3.",
            italic_sm,
        ))
        story.append(Spacer(1, 2 * mm))

        def _kv_table(rows, col_widths=(55 * mm, 115 * mm)):
            data = [[Paragraph(f"<b>{k}</b>", normal_sm), Paragraph(str(v or "--"), normal_sm)]
                    for k, v in rows]
            t = Table(data, colWidths=col_widths)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), GREY_RL),
                ("BACKGROUND", (1, 0), (1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            return t

        # Section 1: Store Details
        story.append(Paragraph("1. Store Details", sec_style))
        store_kv = [(p, v) for p, v in [
            ("Country", country), ("City", city), ("Store Type", store_type),
            ("Area", f"{area_sqft} sq ft" if area_sqft else ""),
            ("Ceiling Height", f"{ceiling_h} ft" if ceiling_h else ""),
            ("Max Temperature", f"{ambient} deg C" if ambient else ""),
            ("Humidity", humidity), ("Budget Level", budget),
        ] if v]
        story.append(_kv_table(store_kv))
        story.append(Spacer(1, 3 * mm))

        # Section 2: HVAC System
        story.append(Paragraph("2. Recommended HVAC System", sec_style))
        sys_kv = [
            ("System Type", system_label),
            ("Capacity", capacity_display),
            ("Selection Basis", selection_basis),
        ]
        if confidence_pct:
            sys_kv.append(("Confidence", f"{confidence_pct}%"))
        sys_kv.append(("Reason / Rationale", rationale))
        story.append(_kv_table(sys_kv))
        story.append(Spacer(1, 3 * mm))

        # Section 3: Scope of Work
        story.append(Paragraph("3. Scope of Work", sec_style))
        hdr = [
            Paragraph("<b>S.No</b>", normal_sm),
            Paragraph("<b>Category</b>", normal_sm),
            Paragraph("<b>Description</b>", normal_sm),
            Paragraph("<b>Unit</b>", normal_sm),
            Paragraph("<b>Qty</b>", normal_sm),
        ]
        scope_pdf_rows = [hdr] + [
            [
                Paragraph(str(i), normal_sm),
                Paragraph(cat, normal_sm),
                Paragraph(desc, normal_sm),
                Paragraph(unit, normal_sm),
                Paragraph(str(qty) if qty != "" else "", normal_sm),
            ]
            for i, (cat, desc, unit, qty) in enumerate(scope_rows, 1)
        ]
        scope_t = Table(scope_pdf_rows, colWidths=[10 * mm, 28 * mm, 80 * mm, 20 * mm, 14 * mm])
        scope_ts = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), LTBLUE_RL),
            ("BACKGROUND", (0, 1), (1, -1), GREY_RL),
            ("BACKGROUND", (2, 1), (-1, -1), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ])
        scope_t.setStyle(scope_ts)
        story.append(scope_t)
        story.append(Spacer(1, 3 * mm))

        # Section 4: Commercial Terms
        story.append(Paragraph("4. Commercial Terms", sec_style))
        comm_kv = [
            ("Quotation Validity", "90 days from submission date"),
            ("Delivery / Completion", "Vendor to specify in quotation"),
            ("Payment Terms", "As per standard purchase order terms"),
            ("Currency", f"{country} local currency or as agreed"),
            ("Pricing Basis", "Lump sum (supply + installation + commissioning)"),
            ("Warranty", "Minimum 2 years on equipment and workmanship"),
            ("Compliance", "ASHRAE 90.1, SASO, and applicable local building code"),
        ]
        story.append(_kv_table(comm_kv))
        story.append(Spacer(1, 3 * mm))

        # Section 5: Vendor Pricing Response
        story.append(Paragraph("5. Vendor Pricing Response (To be filled by Vendor)", sec_style))
        vnd_hdr = [
            Paragraph("<b>S.No</b>", normal_sm),
            Paragraph("<b>Category</b>", normal_sm),
            Paragraph("<b>Description / Proposed Model</b>", normal_sm),
            Paragraph("<b>Unit Price</b>", normal_sm),
            Paragraph("<b>Total</b>", normal_sm),
        ]
        vnd_rows = [vnd_hdr] + [
            [Paragraph(str(i), normal_sm), Paragraph("", normal_sm),
             Paragraph("", normal_sm), Paragraph("", normal_sm), Paragraph("", normal_sm)]
            for i in range(1, len(scope_rows) + 1)
        ] + [[
            Paragraph("", normal_sm),
            Paragraph("", normal_sm),
            Paragraph("<b>GRAND TOTAL (Excl. VAT)</b>", normal_sm),
            Paragraph("", normal_sm),
            Paragraph("", normal_sm),
        ]]
        vnd_t = Table(vnd_rows, colWidths=[10 * mm, 28 * mm, 80 * mm, 20 * mm, 14 * mm])
        vnd_t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), LTBLUE_RL),
            ("BACKGROUND", (-3, -1), (-1, -1), SECBG_RL),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BDBDBD")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(vnd_t)
        story.append(Spacer(1, 4 * mm))

        # Disclaimer
        story.append(Paragraph(
            "<i>Note: This RFQ is system-generated. All specifications are indicative. "
            "A qualified HVAC engineer must review and confirm the final scope before award.</i>",
            ParagraphStyle("disc", parent=styles["Normal"], fontSize=7,
                           textColor=colors.HexColor("#888888")),
        ))

        doc.build(story)
        pdf_buf.seek(0)
        pdf_bytes = pdf_buf.getvalue()
    except Exception as _pdf_exc:
        logger.warning("RFQ PDF build failed: %s", _pdf_exc)
        pdf_bytes = b""

    # ---- Upload both to Azure Blob + persist GeneratedRFQ ----
    xlsx_blob_path = ""
    pdf_blob_path = ""
    if _is_generate_post:
        try:
            from apps.documents.blob_service import upload_to_blob, is_blob_storage_enabled
            if is_blob_storage_enabled():
                _date_str = datetime.date.today().strftime("%Y%m%d")
                _base_name = f"RFQ-{proc_request.pk:04d}-{_date_str}_{safe_title}"
                _folder = f"rfq/{safe_title}"
                if xlsx_bytes:
                    _xblob = f"{_folder}/{_base_name}.xlsx"
                    upload_to_blob(
                        io.BytesIO(xlsx_bytes),
                        _xblob,
                        content_type=(
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        ),
                    )
                    xlsx_blob_path = _xblob
                if pdf_bytes:
                    _pblob = f"{_folder}/{_base_name}.pdf"
                    upload_to_blob(
                        io.BytesIO(pdf_bytes),
                        _pblob,
                        content_type="application/pdf",
                    )
                    pdf_blob_path = _pblob
        except Exception as _blob_exc:
            logger.warning("RFQ blob upload failed: %s", _blob_exc)

        # Persist GeneratedRFQ record
        from apps.procurement.models import GeneratedRFQ as _GRFQ
        _rfq_rec = _GRFQ.objects.create(
            request=proc_request,
            rfq_ref=rfq_ref,
            system_code=system_code,
            system_label=system_label,
            qty_json=qty_overrides,
            xlsx_blob_path=xlsx_blob_path,
            pdf_blob_path=pdf_blob_path,
            generated_by=request.user if request.user.is_authenticated else None,
        )
        return JsonResponse({
            "status": "ok",
            "rfq_ref": rfq_ref,
            "rfq_id": _rfq_rec.pk,
            "has_xlsx": bool(xlsx_bytes),
            "has_pdf": bool(pdf_bytes),
            "blob_enabled": bool(xlsx_blob_path),
        })

    # ---- Legacy GET: stream xlsx (default) or pdf ----
    if _dl_fmt_override == "pdf":
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename_pdf}"'
        return resp

    response = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename_xlsx}"'
    return response


# ---------------------------------------------------------------------------
# 4. Analysis Run Detail
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view_results")
def run_detail(request, pk):
    """Detail view for a single analysis run."""
    run = get_object_or_404(
        AnalysisRun.objects.select_related("request", "triggered_by"),
        pk=pk,
    )

    compliance = None
    validation = None

    if run.run_type == AnalysisRunType.RECOMMENDATION:
        recommendation = RecommendationResult.objects.filter(run=run).first()
    elif run.run_type == AnalysisRunType.BENCHMARK:
        benchmarks = BenchmarkResult.objects.filter(run=run).prefetch_related("lines", "quotation")

    compliance = ComplianceResult.objects.filter(run=run).first()

    # Check for validation result
    from apps.core.enums import AnalysisRunType as ART
    if run.run_type == ART.VALIDATION:
        validation = ValidationResult.objects.filter(run=run).prefetch_related("items").first()

    compliance = ComplianceResult.objects.filter(run=run).first()

    # Audit events for this run
    audit_events = AuditService.fetch_entity_history("AnalysisRun", run.pk)

    return render(request, "procurement/run_detail.html", {
        "run": run,
        "recommendation": recommendation,
        "benchmarks": benchmarks,
        "compliance": compliance,
        "validation": validation,
        "audit_events": audit_events[:30],
    })


# ---------------------------------------------------------------------------
# Actions: Trigger analysis, mark ready, upload quotation
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.run_analysis")
def trigger_analysis(request, pk):
    """Trigger an analysis run from the workspace."""
    if request.method != "POST":
        return redirect("procurement:request_workspace", pk=pk)

    proc_request = get_object_or_404(ProcurementRequest, pk=pk)
    run_type = request.POST.get("run_type", "RECOMMENDATION")

    from apps.procurement.services.analysis_run_service import AnalysisRunService
    from apps.procurement.tasks import run_analysis_task

    run = AnalysisRunService.create_run(
        request=proc_request,
        run_type=run_type,
        triggered_by=request.user,
    )
    run_analysis_task.delay(run.pk)
    messages.success(request, f"{run_type} analysis queued (Run {run.run_id}).")
    return redirect("procurement:request_workspace", pk=pk)


@login_required
@permission_required_code("procurement.edit")
def mark_ready(request, pk):
    """Mark a request as READY after attribute validation."""
    if request.method != "POST":
        return redirect("procurement:request_workspace", pk=pk)

    proc_request = get_object_or_404(ProcurementRequest, pk=pk)
    from apps.procurement.services.request_service import ProcurementRequestService
    try:
        ProcurementRequestService.mark_ready(proc_request, user=request.user)
        messages.success(request, "Request marked as ready.")
    except ValueError as exc:
        messages.error(request, str(exc))
    return redirect("procurement:request_workspace", pk=pk)


@login_required
@permission_required_code("procurement.manage_quotations")
def upload_quotation(request, pk):
    """Upload a supplier quotation to a request."""
    if request.method != "POST":
        return redirect("procurement:request_workspace", pk=pk)

    proc_request = get_object_or_404(ProcurementRequest, pk=pk)
    from apps.procurement.services.quotation_service import QuotationService

    try:
        quotation = QuotationService.create_quotation(
            request=proc_request,
            vendor_name=request.POST.get("vendor_name", ""),
            quotation_number=request.POST.get("quotation_number", ""),
            total_amount=request.POST.get("total_amount") or None,
            currency=request.POST.get("currency", "USD"),
            created_by=request.user,
        )
        messages.success(request, f"Quotation from '{quotation.vendor_name}' added.")
    except Exception as exc:
        messages.error(request, f"Failed to add quotation: {exc}")
    return redirect("procurement:request_workspace", pk=pk)


# ---------------------------------------------------------------------------
# Trigger Validation
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.validate")
def trigger_validation(request, pk):
    """Trigger a validation run from the workspace."""
    if request.method != "POST":
        return redirect("procurement:request_workspace", pk=pk)

    proc_request = get_object_or_404(ProcurementRequest, pk=pk)

    from apps.core.enums import AnalysisRunType
    from apps.procurement.services.analysis_run_service import AnalysisRunService
    from apps.procurement.tasks import run_validation_task

    run = AnalysisRunService.create_run(
        request=proc_request,
        run_type=AnalysisRunType.VALIDATION,
        triggered_by=request.user,
    )
    run_validation_task.delay(run.pk)
    messages.success(request, f"Validation queued (Run {run.run_id}).")
    return redirect("procurement:request_workspace", pk=pk)


# ---------------------------------------------------------------------------
# Quotation Prefill Review
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.manage_quotations")
def quotation_prefill_review(request, pk):
    """Review and confirm extracted quotation data from PDF prefill."""
    import json
    quotation = get_object_or_404(
        SupplierQuotation.objects.select_related("request", "uploaded_document"),
        pk=pk,
    )
    payload = quotation.prefill_payload_json or {}
    return render(request, "procurement/quotation_prefill_review.html", {
        "quotation": quotation,
        "payload": payload,
        "payload_json": json.dumps(payload),
    })


# ---------------------------------------------------------------------------
# Procurement Manager Dashboard — Landmark Group
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view")
def procurement_dashboard(request):
    """Landmark Group Procurement Manager Dashboard with KPIs, charts, and quick actions."""
    from django.db.models import Count

    qs = ProcurementRequest.objects.all()

    # KPI counts
    status_counts = {item["status"]: item["total"] for item in qs.values("status").annotate(total=Count("id"))}
    kpi = {
        "total": qs.count(),
        "draft": status_counts.get("DRAFT", 0),
        "ready": status_counts.get("READY", 0),
        "pending_approval": status_counts.get("REVIEW_REQUIRED", 0),
        "approved": status_counts.get("COMPLETED", 0),
        "hvac": qs.filter(domain_code="HVAC").count(),
    }

    # Status chart data
    status_chart = list(qs.values("status").annotate(total=Count("id")).order_by("-total"))

    # Domain breakdown chart
    domain_chart = list(
        qs.exclude(domain_code="").values("domain_code").annotate(total=Count("id")).order_by("-total")[:10]
    )

    # GCC country breakdown
    from django.db.models import F as _F
    by_country = list(
        qs.exclude(geography_country__isnull=True).exclude(geography_country="")
        .values(country=_F("geography_country")).annotate(total=Count("id")).order_by("-total")
    )

    # HVAC type breakdown using attributes
    hvac_type_labels = {
        "SPLIT_AC": "Split AC",
        "CASSETTE_AC": "Cassette AC",
        "VRF_VRV": "VRF/VRV System",
        "FCU_CW": "FCU / Chilled Water",
        "AHU": "Air Handling Unit",
        "PACKAGED_UNIT": "Packaged Unit",
        "CHILLER": "Chiller Plant",
        "VENTILATION": "Ventilation / MEP",
    }
    from apps.procurement.models import ProcurementRequestAttribute
    hvac_attr_qs = (
        ProcurementRequestAttribute.objects.filter(
            request__domain_code="HVAC",
            attribute_code="product_type",
        )
        .values("value_text")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    hvac_by_type = [
        {"type_label": hvac_type_labels.get(row["value_text"] or "", row["value_text"] or "Unknown"), "total": row["total"]}
        for row in hvac_attr_qs
    ]

    # Recent requests
    recent_requests = (
        qs.select_related("created_by")
        .order_by("-created_at")[:10]
    )

    # ------------------------------------------------------------------
    # Benchmarking KPIs
    # ------------------------------------------------------------------
    from apps.benchmarking.models import BenchmarkRequest, BenchmarkResult

    bq_total = BenchmarkRequest.objects.count()
    bq_completed = BenchmarkRequest.objects.filter(status="COMPLETED").count()
    bq_pending = BenchmarkRequest.objects.filter(status__in=["PENDING", "PROCESSING"]).count()
    bq_failed = BenchmarkRequest.objects.filter(status="FAILED").count()

    # Variance distribution across all completed results
    from django.db.models import Sum
    variance_agg = BenchmarkResult.objects.aggregate(
        within=Sum("lines_within_range"),
        moderate=Sum("lines_moderate"),
        high=Sum("lines_high"),
        needs_review=Sum("lines_needs_review"),
    )
    bq_kpi = {
        "total": bq_total,
        "completed": bq_completed,
        "pending": bq_pending,
        "failed": bq_failed,
        "lines_within_range": variance_agg["within"] or 0,
        "lines_moderate": variance_agg["moderate"] or 0,
        "lines_high": variance_agg["high"] or 0,
        "lines_needs_review": variance_agg["needs_review"] or 0,
    }

    # Geography breakdown for benchmark requests
    bench_by_geo = list(
        BenchmarkRequest.objects
        .exclude(geography="")
        .values("geography")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # Recent benchmark requests with their result
    recent_benchmarks = (
        BenchmarkRequest.objects
        .select_related("submitted_by", "result")
        .order_by("-created_at")[:10]
    )

    return render(request, "procurement/procurement_dashboard.html", {
        "kpi": kpi,
        "status_chart": status_chart,
        "domain_chart": domain_chart,
        "by_country": by_country,
        "hvac_by_type": hvac_by_type,
        "recent_requests": recent_requests,
        # benchmarking
        "bq_kpi": bq_kpi,
        "bench_by_geo": bench_by_geo,
        "recent_benchmarks": recent_benchmarks,
    })


# ===========================================================================
# HVAC FLOW A -- new dedicated views
# ===========================================================================

# ---------------------------------------------------------------------------
# H1. HVAC Request List (dashboard)
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view")
def hvac_request_list(request):
    """List all HVAC procurement requests with KPIs, search, and filters."""
    base_qs = ProcurementRequest.objects.filter(domain_code="HVAC")

    # KPI totals
    total = base_qs.count()
    completed = base_qs.filter(status="COMPLETED").count()
    in_progress = base_qs.filter(status__in=["IN_PROGRESS", "RUNNING", "READY"]).count()
    # needs_review: requests whose latest recommendation has human_review flags
    needs_review = base_qs.filter(status="REVIEW_REQUIRED").count()

    # Annotate with latest recommendation data
    latest_rec_option_sq = Subquery(
        RecommendationResult.objects.filter(run__request=OuterRef("pk"))
        .order_by("-created_at")
        .values("recommended_option")[:1]
    )
    latest_conf_sq = Subquery(
        RecommendationResult.objects.filter(run__request=OuterRef("pk"))
        .order_by("-created_at")
        .values("confidence_score")[:1]
    )

    qs = base_qs.select_related("created_by").annotate(
        run_count=Count("analysis_runs"),
        latest_recommended_option=latest_rec_option_sq,
        latest_confidence_score=latest_conf_sq,
    )

    # Filters
    status_filter = request.GET.get("status", "")
    search = request.GET.get("q", "")
    sort = request.GET.get("sort", "newest")

    if status_filter:
        qs = qs.filter(status=status_filter)
    if search:
        qs = qs.filter(
            Q(title__icontains=search)
            | Q(description__icontains=search)
            | Q(geography_country__icontains=search)
            | Q(geography_city__icontains=search)
        )
    if sort == "oldest":
        qs = qs.order_by("created_at")
    elif sort == "confidence":
        qs = qs.order_by("-latest_confidence_score")
    else:
        qs = qs.order_by("-created_at")

    paginator = Paginator(qs, 12)
    page = paginator.get_page(request.GET.get("page"))

    return render(request, "procurement/hvac_request_list.html", {
        "page_obj": page,
        "kpi": {
            "total": total,
            "completed": completed,
            "in_progress": in_progress,
            "needs_review": needs_review,
        },
        "current_status": status_filter,
        "search_query": search,
        "current_sort": sort,
        "status_choices": ProcurementRequestStatus.choices,
    })


# ---------------------------------------------------------------------------
# H2. HVAC Request Detail (recommendation workspace)
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view")
def hvac_request_detail(request, pk):
    """Full recommendation workspace for a single HVAC request."""
    proc_request = get_object_or_404(
        ProcurementRequest.objects.select_related("created_by", "assigned_to"),
        pk=pk, domain_code="HVAC",
    )
    attributes = proc_request.attributes.order_by("attribute_code")
    runs = proc_request.analysis_runs.select_related("triggered_by").order_by("-created_at")
    latest_run = runs.first()

    recommendation = (
        RecommendationResult.objects.filter(run__request=proc_request)
        .select_related("run")
        .order_by("-created_at")
        .first()
    )

    # Parse structured payload sections safely
    payload = {}
    if recommendation and recommendation.output_payload_json:
        raw = recommendation.output_payload_json
        if isinstance(raw, str):
            try:
                payload = json.loads(raw)
            except Exception:
                payload = {}
        elif isinstance(raw, dict):
            payload = raw

    # Agent execution records for this request (all runs)
    from apps.procurement.models import ProcurementAgentExecutionRecord
    agent_records = ProcurementAgentExecutionRecord.objects.filter(
        analysis_run__request=proc_request
    ).order_by("-created_at")[:20]

    # Latest compliance
    compliance = (
        ComplianceResult.objects.filter(run__request=proc_request)
        .select_related("run")
        .order_by("-created_at")
        .first()
    )

    # Audit events
    audit_events = AuditService.fetch_entity_history("ProcurementRequest", proc_request.pk)

    # Benchmarks
    benchmarks = BenchmarkResult.objects.filter(
        run__request=proc_request,
    ).select_related("run", "quotation").prefetch_related("lines").order_by("-created_at")

    return render(request, "procurement/hvac_request_detail.html", {
        "proc_request": proc_request,
        "attributes": attributes,
        "runs": runs,
        "latest_run": latest_run,
        "recommendation": recommendation,
        "payload": payload,
        "agent_records": agent_records,
        "compliance": compliance,
        "benchmarks": benchmarks,
        "audit_events": audit_events[:30],
    })


# ---------------------------------------------------------------------------
# H3. HVAC Request Form (create new)
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.create")
def hvac_request_form(request):
    """Create a new HVAC procurement request via the structured HVAC form."""
    if request.method == "POST":
        from apps.procurement.services.request_service import ProcurementRequestService
        from apps.procurement.services.analysis_run_service import AnalysisRunService
        from apps.procurement.tasks import run_analysis_task

        # Map flat form fields to ProcurementRequestAttribute records
        HVAC_FIELD_MAP = [
            ("store_id", "Store ID / Reference", "TEXT"),
            ("brand", "Brand / Operator", "TEXT"),
            ("geography_zone", "Geography Zone", "TEXT"),
            ("store_type", "Store Type", "TEXT"),
            ("store_format", "Store Format", "TEXT"),
            ("area_sqft", "Store Area (sqft)", "NUMBER"),
            ("ceiling_height_ft", "Ceiling Height (ft)", "NUMBER"),
            ("ambient_temp_max", "Ambient Temp Max (C)", "NUMBER"),
            ("humidity_level", "Humidity Level", "TEXT"),
            ("dust_exposure", "Dust Exposure", "TEXT"),
            ("heat_load_category", "Heat Load Category", "TEXT"),
            ("fresh_air_requirement", "Fresh Air Requirement", "TEXT"),
            ("existing_hvac_type", "Existing HVAC Type", "TEXT"),
            ("landlord_constraints", "Landlord Constraints", "TEXT"),
            ("required_standards_notes", "Required Standards / Notes", "TEXT"),
            ("budget_level", "Budget Level", "TEXT"),
            ("energy_efficiency_priority", "Energy Efficiency Priority", "TEXT"),
            ("maintenance_priority", "Maintenance Priority", "TEXT"),
            ("preferred_oems", "Preferred OEMs", "TEXT"),
        ]

        attrs_data = []
        for code, label, dtype in HVAC_FIELD_MAP:
            val = request.POST.get(code, "").strip()
            if val:
                attrs_data.append({
                    "attribute_code": code,
                    "attribute_label": label,
                    "data_type": dtype,
                    "value_text": val,
                    "is_required": False,
                })

        try:
            proc_request = ProcurementRequestService.create_request(
                title=request.POST.get("title", ""),
                description=request.POST.get("description", ""),
                domain_code="HVAC",
                schema_code="HVAC_FLOW_A",
                request_type=ProcurementRequestType.RECOMMENDATION,
                priority=request.POST.get("priority", "HIGH"),
                geography_country=request.POST.get("geography_country", "UAE"),
                geography_city=request.POST.get("geography_city", ""),
                currency="AED",
                created_by=request.user,
                attributes=attrs_data if attrs_data else None,
            )
            action = request.POST.get("action", "save")
            if action == "run":
                run = AnalysisRunService.create_run(
                    request=proc_request,
                    run_type=AnalysisRunType.RECOMMENDATION,
                    triggered_by=request.user,
                )
                run_analysis_task.delay(run.pk)
                messages.success(request, "HVAC request created and analysis queued.")
            else:
                messages.success(request, f"HVAC request '{proc_request.title}' saved.")
            return redirect("procurement:hvac_request_detail", pk=proc_request.pk)
        except Exception as exc:
            logger.exception("hvac_request_form POST failed: %s", exc)
            messages.error(request, f"Failed to save request: {exc}")

    return render(request, "procurement/hvac_request_form.html", {
        "priority_choices": [("LOW", "Low"), ("MEDIUM", "Medium"), ("HIGH", "High"), ("CRITICAL", "Critical")],
    })


# ---------------------------------------------------------------------------
# H4. Benchmarking List
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view")
def hvac_benchmark_list(request):
    """List all HVAC benchmarking runs."""
    qs = BenchmarkResult.objects.filter(
        run__request__domain_code="HVAC",
    ).select_related("run", "run__request", "quotation").prefetch_related("lines").order_by("-created_at")

    total = qs.count()
    completed = qs.filter(run__status="COMPLETED").count()

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get("page"))

    return render(request, "procurement/hvac_benchmark_list.html", {
        "page_obj": page,
        "kpi": {"total": total, "completed": completed},
    })


# ---------------------------------------------------------------------------
# H5. HVAC Configuration
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.edit")
def hvac_config(request):
    """HVAC Flow A configuration -- source registry and rule engine settings."""
    if request.method == "POST":
        action = request.POST.get("config_action", "")
        if action == "add_source":
            try:
                ExternalSourceRegistry.objects.create(
                    source_name=request.POST.get("source_name", ""),
                    domain=request.POST.get("domain", ""),
                    source_type=request.POST.get("source_type", ExternalSourceClass.OEM_OFFICIAL),
                    country_scope=[c.strip() for c in request.POST.get("country_scope", "").split(",") if c.strip()],
                    priority=int(request.POST.get("priority", 10)),
                    trust_score=float(request.POST.get("trust_score", 0.8)),
                    allowed_for_discovery=request.POST.get("allowed_for_discovery") == "on",
                    allowed_for_compliance=request.POST.get("allowed_for_compliance") == "on",
                    fetch_mode=request.POST.get("fetch_mode", "PAGE"),
                    notes=request.POST.get("notes", ""),
                )
                messages.success(request, "External source added.")
            except Exception as exc:
                logger.exception("add_source failed: %s", exc)
                messages.error(request, f"Failed to add source: {exc}")
        elif action == "toggle_source":
            src_id = request.POST.get("source_id")
            if src_id:
                try:
                    src = ExternalSourceRegistry.objects.get(pk=src_id)
                    src.is_active = not src.is_active
                    src.save(update_fields=["is_active"])
                    messages.success(request, f"Source '{src.source_name}' toggled.")
                except ExternalSourceRegistry.DoesNotExist:
                    messages.error(request, "Source not found.")
        return redirect("procurement:hvac_config")

    sources = ExternalSourceRegistry.objects.all().order_by("priority", "source_name")

    return render(request, "procurement/hvac_config.html", {
        "sources": sources,
        "source_type_choices": ExternalSourceClass.choices,
        "fetch_mode_choices": [("PAGE", "Web Page"), ("PDF", "PDF Download"), ("API", "API Endpoint")],
    })


# =============================================================================
# PROCUREMENT CONFIGURATIONS -- full admin CRUD hub (AJAX-backed)
# =============================================================================

@login_required
@permission_required_code("procurement.view")
def proc_configurations(request):
    """Procurement Configurations hub -- seed data admin control."""
    from apps.procurement.models import (
        ExternalSourceRegistry, ValidationRuleSet, Product, Vendor, Room,
        HVACRecommendationRule, HVACServiceScope,
    )
    from apps.core.enums import (
        ExternalSourceClass, ValidationType, HVACSystemType, RoomUsageType,
    )

    stats = {
        "sources_total": ExternalSourceRegistry.objects.count(),
        "sources_active": ExternalSourceRegistry.objects.filter(is_active=True).count(),
        "rulesets_total": ValidationRuleSet.objects.count(),
        "rulesets_active": ValidationRuleSet.objects.filter(is_active=True).count(),
        "products_total": Product.objects.count(),
        "products_active": Product.objects.filter(is_active=True).count(),
        "vendors_total": Vendor.objects.count(),
        "vendors_active": Vendor.objects.filter(is_active=True).count(),
        "rooms_total": Room.objects.count(),
        "rooms_active": Room.objects.filter(is_active=True).count(),
        "hvacrules_total": HVACRecommendationRule.objects.count(),
        "hvacrules_active": HVACRecommendationRule.objects.filter(is_active=True).count(),
        "service_total": HVACServiceScope.objects.count(),
        "service_active": HVACServiceScope.objects.filter(is_active=True).count(),
    }

    return render(request, "procurement/configurations.html", {
        "stats": stats,
        "source_type_choices": ExternalSourceClass.choices,
        "fetch_mode_choices": [("PAGE", "Web Page"), ("PDF", "PDF Download"), ("API", "API Endpoint")],
        "validation_type_choices": ValidationType.choices,
        "system_type_choices": HVACSystemType.choices,
        "room_usage_choices": RoomUsageType.choices,
        "active_tab": request.GET.get("tab", "sources"),
        # HVAC rules filter choices
        "hvac_store_type_choices": HVACRecommendationRule.STORE_TYPE_CHOICES,
        "hvac_budget_choices": HVACRecommendationRule.BUDGET_CHOICES,
        "hvac_energy_priority_choices": HVACRecommendationRule.ENERGY_PRIORITY_CHOICES,
    })


# ---------------------------------------------------------------------------
# AJAX API -- External Source Registry
# ---------------------------------------------------------------------------

@login_required
@permission_required_code("procurement.view")
def api_config_sources(request):
    """List (GET) or Create (POST) ExternalSourceRegistry entries."""
    from apps.procurement.models import ExternalSourceRegistry
    from django.http import JsonResponse

    if request.method == "GET":
        q = request.GET.get("q", "").strip()
        qs = ExternalSourceRegistry.objects.all().order_by("priority", "source_name")
        if q:
            qs = qs.filter(
                Q(source_name__icontains=q) | Q(domain__icontains=q) | Q(source_type__icontains=q)
            )
        items = [
            {
                "id": s.pk,
                "source_name": s.source_name,
                "domain": s.domain,
                "source_url": s.source_url or "",
                "hvac_system_type": s.hvac_system_type,
                "equipment": s.equipment,
                "source_type": s.source_type,
                "source_type_display": s.get_source_type_display(),
                "priority": s.priority,
                "trust_score": float(s.trust_score),
                "allowed_for_discovery": s.allowed_for_discovery,
                "allowed_for_compliance": s.allowed_for_compliance,
                "fetch_mode": s.fetch_mode,
                "country_scope": ", ".join(s.country_scope) if s.country_scope else "",
                "notes": s.notes,
                "is_active": s.is_active,
            }
            for s in qs
        ]
        return JsonResponse({"items": items, "total": len(items)})

    if request.method == "POST":
        if not request.user.has_perm("procurement.edit") and not request.user.is_staff:
            # Simple permission check -- defer to RBAC
            pass
        try:
            body = json.loads(request.body)
            src = ExternalSourceRegistry.objects.create(
                source_name=body.get("source_name", "").strip(),
                domain=body.get("domain", "").strip(),
                hvac_system_type=body.get("hvac_system_type", "").strip(),
                equipment=body.get("equipment", "").strip(),
                source_url=body.get("source_url", "").strip(),
                source_type=body.get("source_type", "OEM_OFFICIAL"),
                country_scope=[c.strip() for c in body.get("country_scope", "").split(",") if c.strip()],
                priority=int(body.get("priority", 10)),
                trust_score=float(body.get("trust_score", 0.8)),
                allowed_for_discovery=bool(body.get("allowed_for_discovery", True)),
                allowed_for_compliance=bool(body.get("allowed_for_compliance", False)),
                fetch_mode=body.get("fetch_mode", "PAGE"),
                notes=body.get("notes", ""),
                is_active=bool(body.get("is_active", True)),
            )
            return JsonResponse({
                "success": True, "id": src.pk,
                "message": f"Source '{src.source_name}' created successfully.",
            })
        except Exception as exc:
            logger.exception("api_config_sources POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
@permission_required_code("procurement.view")
def api_config_source_detail(request, pk):
    """Get (GET), Update (PUT-via-POST) or Delete (DELETE-via-POST) a source."""
    from apps.procurement.models import ExternalSourceRegistry
    from django.http import JsonResponse

    try:
        src = ExternalSourceRegistry.objects.get(pk=pk)
    except ExternalSourceRegistry.DoesNotExist:
        return JsonResponse({"success": False, "message": "Source not found."}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": src.pk,
            "source_name": src.source_name,
            "domain": src.domain,
            "hvac_system_type": src.hvac_system_type,
            "equipment": src.equipment,
            "source_url": src.source_url or "",
            "source_type": src.source_type,
            "priority": src.priority,
            "trust_score": float(src.trust_score),
            "allowed_for_discovery": src.allowed_for_discovery,
            "allowed_for_compliance": src.allowed_for_compliance,
            "fetch_mode": src.fetch_mode,
            "country_scope": ", ".join(src.country_scope) if src.country_scope else "",
            "notes": src.notes,
            "is_active": src.is_active,
        })

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            action = body.get("_action", "update")

            if action == "delete":
                name = src.source_name
                src.delete()
                return JsonResponse({"success": True, "message": f"Source '{name}' deleted."})

            if action == "toggle":
                src.is_active = not src.is_active
                src.save(update_fields=["is_active"])
                return JsonResponse({"success": True, "is_active": src.is_active,
                                     "message": f"Source '{src.source_name}' {'activated' if src.is_active else 'deactivated'}."})

            # update
            src.source_name = body.get("source_name", src.source_name).strip()
            src.domain = body.get("domain", src.domain).strip()
            src.hvac_system_type = body.get("hvac_system_type", src.hvac_system_type or "").strip()
            src.equipment = body.get("equipment", src.equipment or "").strip()
            src.source_url = body.get("source_url", src.source_url or "").strip()
            src.source_type = body.get("source_type", src.source_type)
            country_raw = body.get("country_scope", "")
            src.country_scope = [c.strip() for c in country_raw.split(",") if c.strip()] if country_raw else src.country_scope
            src.priority = int(body.get("priority", src.priority))
            src.trust_score = float(body.get("trust_score", src.trust_score))
            src.allowed_for_discovery = bool(body.get("allowed_for_discovery", src.allowed_for_discovery))
            src.allowed_for_compliance = bool(body.get("allowed_for_compliance", src.allowed_for_compliance))
            src.fetch_mode = body.get("fetch_mode", src.fetch_mode)
            src.notes = body.get("notes", src.notes)
            src.is_active = bool(body.get("is_active", src.is_active))
            src.save()
            return JsonResponse({"success": True, "message": f"Source '{src.source_name}' updated."})
        except Exception as exc:
            logger.exception("api_config_source_detail POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
@permission_required_code("procurement.view")
def api_config_validate_url(request):
    """HEAD-check a URL and return whether it is reachable (no broken link)."""
    import requests as _req
    from django.http import JsonResponse

    url = request.GET.get("url", "").strip()
    if not url:
        return JsonResponse({"ok": False, "error": "No URL provided."})
    if not url.startswith(("http://", "https://")):
        return JsonResponse({"ok": False, "error": "URL must start with http:// or https://"})
    try:
        resp = _req.head(
            url,
            timeout=8,
            allow_redirects=True,
            headers={"User-Agent": "Mozilla/5.0 (compatible; LinkChecker/1.0)"},
        )
        ok = resp.status_code < 400
        return JsonResponse({"ok": ok, "status_code": resp.status_code})
    except _req.exceptions.Timeout:
        return JsonResponse({"ok": False, "error": "Request timed out."})
    except _req.exceptions.ConnectionError:
        return JsonResponse({"ok": False, "error": "Could not connect to the server."})
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)})


# ---------------------------------------------------------------------------
# AJAX API -- Validation Rule Sets
# ---------------------------------------------------------------------------

@login_required
@permission_required_code("procurement.view")
def api_config_rulesets(request):
    """List (GET) or Create (POST) ValidationRuleSet entries."""
    from apps.procurement.models import ValidationRuleSet
    from django.http import JsonResponse

    if request.method == "GET":
        q = request.GET.get("q", "").strip()
        qs = ValidationRuleSet.objects.all().order_by("priority", "rule_set_code")
        if q:
            qs = qs.filter(
                Q(rule_set_code__icontains=q) | Q(rule_set_name__icontains=q) | Q(domain_code__icontains=q)
            )
        items = [
            {
                "id": r.pk,
                "rule_set_code": r.rule_set_code,
                "rule_set_name": r.rule_set_name,
                "domain_code": r.domain_code,
                "schema_code": r.schema_code,
                "validation_type": r.validation_type,
                "validation_type_display": r.get_validation_type_display(),
                "priority": r.priority,
                "description": r.description,
                "is_active": r.is_active,
            }
            for r in qs
        ]
        return JsonResponse({"items": items, "total": len(items)})

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            rs = ValidationRuleSet.objects.create(
                rule_set_code=body.get("rule_set_code", "").strip(),
                rule_set_name=body.get("rule_set_name", "").strip(),
                domain_code=body.get("domain_code", "").strip(),
                schema_code=body.get("schema_code", "").strip(),
                validation_type=body.get("validation_type", "ATTRIBUTE_COMPLETENESS"),
                priority=int(body.get("priority", 100)),
                description=body.get("description", ""),
                is_active=bool(body.get("is_active", True)),
            )
            return JsonResponse({"success": True, "id": rs.pk,
                                 "message": f"Rule set '{rs.rule_set_code}' created."})
        except Exception as exc:
            logger.exception("api_config_rulesets POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
@permission_required_code("procurement.view")
def api_config_ruleset_detail(request, pk):
    """Get, Update, or Delete a ValidationRuleSet."""
    from apps.procurement.models import ValidationRuleSet
    from django.http import JsonResponse

    try:
        rs = ValidationRuleSet.objects.get(pk=pk)
    except ValidationRuleSet.DoesNotExist:
        return JsonResponse({"success": False, "message": "Rule set not found."}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": rs.pk,
            "rule_set_code": rs.rule_set_code,
            "rule_set_name": rs.rule_set_name,
            "domain_code": rs.domain_code,
            "schema_code": rs.schema_code,
            "validation_type": rs.validation_type,
            "priority": rs.priority,
            "description": rs.description,
            "is_active": rs.is_active,
        })

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            action = body.get("_action", "update")

            if action == "delete":
                code = rs.rule_set_code
                rs.delete()
                return JsonResponse({"success": True, "message": f"Rule set '{code}' deleted."})

            if action == "toggle":
                rs.is_active = not rs.is_active
                rs.save(update_fields=["is_active"])
                return JsonResponse({"success": True, "is_active": rs.is_active,
                                     "message": f"Rule set '{rs.rule_set_code}' {'activated' if rs.is_active else 'deactivated'}."})

            rs.rule_set_code = body.get("rule_set_code", rs.rule_set_code).strip()
            rs.rule_set_name = body.get("rule_set_name", rs.rule_set_name).strip()
            rs.domain_code = body.get("domain_code", rs.domain_code).strip()
            rs.schema_code = body.get("schema_code", rs.schema_code).strip()
            rs.validation_type = body.get("validation_type", rs.validation_type)
            rs.priority = int(body.get("priority", rs.priority))
            rs.description = body.get("description", rs.description)
            rs.is_active = bool(body.get("is_active", rs.is_active))
            rs.save()
            return JsonResponse({"success": True, "message": f"Rule set '{rs.rule_set_code}' updated."})
        except Exception as exc:
            logger.exception("api_config_ruleset_detail POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ---------------------------------------------------------------------------
# AJAX API -- Products (HVAC catalog)
# ---------------------------------------------------------------------------

@login_required
@permission_required_code("procurement.view")
def api_config_products(request):
    """List (GET) or Create (POST) Product entries."""
    from apps.procurement.models import Product
    from django.http import JsonResponse

    if request.method == "GET":
        q = request.GET.get("q", "").strip()
        qs = Product.objects.all().order_by("manufacturer", "system_type", "capacity_kw")
        if q:
            qs = qs.filter(
                Q(manufacturer__icontains=q) | Q(product_name__icontains=q) | Q(sku__icontains=q)
            )
        items = [
            {
                "id": p.pk,
                "sku": p.sku,
                "manufacturer": p.manufacturer,
                "product_name": p.product_name,
                "system_type": p.system_type,
                "system_type_display": p.get_system_type_display(),
                "capacity_kw": float(p.capacity_kw),
                "power_input_kw": float(p.power_input_kw),
                "sound_level_db_full_load": p.sound_level_db_full_load,
                "refrigerant_type": p.refrigerant_type,
                "warranty_months": p.warranty_months,
                "cop_rating": float(p.cop_rating) if p.cop_rating is not None else None,
                "seer_rating": float(p.seer_rating) if p.seer_rating is not None else None,
                "weight_kg": p.weight_kg,
                "is_active": p.is_active,
            }
            for p in qs
        ]
        return JsonResponse({"items": items, "total": len(items)})

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            prod = Product.objects.create(
                sku=body.get("sku", "").strip(),
                manufacturer=body.get("manufacturer", "").strip(),
                product_name=body.get("product_name", "").strip(),
                system_type=body.get("system_type", "SPLIT_AC"),
                capacity_kw=float(body.get("capacity_kw", 0)),
                power_input_kw=float(body.get("power_input_kw", 0)),
                sound_level_db_full_load=int(body.get("sound_level_db_full_load", 50)),
                sound_level_db_part_load=int(body.get("sound_level_db_part_load") or 0) or None,
                refrigerant_type=body.get("refrigerant_type", ""),
                warranty_months=int(body.get("warranty_months", 12)),
                cop_rating=float(body.get("cop_rating")) if body.get("cop_rating") else None,
                seer_rating=float(body.get("seer_rating")) if body.get("seer_rating") else None,
                weight_kg=int(body.get("weight_kg")) if body.get("weight_kg") else None,
                installation_support_required=bool(body.get("installation_support_required", False)),
                is_active=bool(body.get("is_active", True)),
            )
            return JsonResponse({"success": True, "id": prod.pk,
                                 "message": f"Product '{prod.sku}' created."})
        except Exception as exc:
            logger.exception("api_config_products POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
@permission_required_code("procurement.view")
def api_config_product_detail(request, pk):
    """Get, Update, or Delete a Product."""
    from apps.procurement.models import Product
    from django.http import JsonResponse

    try:
        prod = Product.objects.get(pk=pk)
    except Product.DoesNotExist:
        return JsonResponse({"success": False, "message": "Product not found."}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": prod.pk,
            "sku": prod.sku,
            "manufacturer": prod.manufacturer,
            "product_name": prod.product_name,
            "system_type": prod.system_type,
            "capacity_kw": float(prod.capacity_kw),
            "power_input_kw": float(prod.power_input_kw),
            "sound_level_db_full_load": prod.sound_level_db_full_load,
            "sound_level_db_part_load": prod.sound_level_db_part_load,
            "refrigerant_type": prod.refrigerant_type,
            "warranty_months": prod.warranty_months,
            "cop_rating": float(prod.cop_rating) if prod.cop_rating is not None else "",
            "seer_rating": float(prod.seer_rating) if prod.seer_rating is not None else "",
            "weight_kg": prod.weight_kg or "",
            "installation_support_required": prod.installation_support_required,
            "is_active": prod.is_active,
        })

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            action = body.get("_action", "update")

            if action == "delete":
                name = prod.sku
                prod.delete()
                return JsonResponse({"success": True, "message": f"Product '{name}' deleted."})

            if action == "toggle":
                prod.is_active = not prod.is_active
                prod.save(update_fields=["is_active"])
                return JsonResponse({"success": True, "is_active": prod.is_active,
                                     "message": f"Product '{prod.sku}' {'activated' if prod.is_active else 'deactivated'}."})

            prod.sku = body.get("sku", prod.sku).strip()
            prod.manufacturer = body.get("manufacturer", prod.manufacturer).strip()
            prod.product_name = body.get("product_name", prod.product_name).strip()
            prod.system_type = body.get("system_type", prod.system_type)
            prod.capacity_kw = float(body.get("capacity_kw", prod.capacity_kw))
            prod.power_input_kw = float(body.get("power_input_kw", prod.power_input_kw))
            prod.sound_level_db_full_load = int(body.get("sound_level_db_full_load", prod.sound_level_db_full_load))
            snd_part = body.get("sound_level_db_part_load")
            prod.sound_level_db_part_load = int(snd_part) if snd_part else None
            prod.refrigerant_type = body.get("refrigerant_type", prod.refrigerant_type)
            prod.warranty_months = int(body.get("warranty_months", prod.warranty_months))
            cop = body.get("cop_rating")
            prod.cop_rating = float(cop) if cop else None
            seer = body.get("seer_rating")
            prod.seer_rating = float(seer) if seer else None
            wt = body.get("weight_kg")
            prod.weight_kg = int(wt) if wt else None
            prod.installation_support_required = bool(body.get("installation_support_required", prod.installation_support_required))
            prod.is_active = bool(body.get("is_active", prod.is_active))
            prod.save()
            return JsonResponse({"success": True, "message": f"Product '{prod.sku}' updated."})
        except Exception as exc:
            logger.exception("api_config_product_detail POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ---------------------------------------------------------------------------
# AJAX API -- Vendors (HVAC vendor master)
# ---------------------------------------------------------------------------

@login_required
@permission_required_code("procurement.view")
def api_config_vendors(request):
    """List (GET) or Create (POST) Vendor entries."""
    from apps.procurement.models import Vendor
    from django.http import JsonResponse

    if request.method == "GET":
        q = request.GET.get("q", "").strip()
        qs = Vendor.objects.all().order_by("vendor_name")
        if q:
            qs = qs.filter(
                Q(vendor_name__icontains=q) | Q(country__icontains=q) | Q(city__icontains=q)
            )
        items = [
            {
                "id": v.pk,
                "vendor_name": v.vendor_name,
                "country": v.country,
                "city": v.city,
                "contact_email": v.contact_email,
                "contact_phone": v.contact_phone,
                "preferred_vendor": v.preferred_vendor,
                "reliability_score": float(v.reliability_score),
                "average_lead_time_days": v.average_lead_time_days,
                "payment_terms": v.payment_terms,
                "bulk_discount_available": v.bulk_discount_available,
                "rush_order_capable": v.rush_order_capable,
                "on_time_delivery_pct": float(v.on_time_delivery_pct),
                "total_purchases": v.total_purchases,
                "notes": v.notes,
                "is_active": v.is_active,
            }
            for v in qs
        ]
        return JsonResponse({"items": items, "total": len(items)})

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            vend = Vendor.objects.create(
                vendor_name=body.get("vendor_name", "").strip(),
                country=body.get("country", "").strip(),
                city=body.get("city", "").strip(),
                address=body.get("address", ""),
                contact_email=body.get("contact_email", ""),
                contact_phone=body.get("contact_phone", ""),
                preferred_vendor=bool(body.get("preferred_vendor", False)),
                reliability_score=float(body.get("reliability_score", 3.5)),
                average_lead_time_days=int(body.get("average_lead_time_days", 7)),
                payment_terms=body.get("payment_terms", ""),
                bulk_discount_available=bool(body.get("bulk_discount_available", False)),
                rush_order_capable=bool(body.get("rush_order_capable", False)),
                on_time_delivery_pct=float(body.get("on_time_delivery_pct", 95.0)),
                notes=body.get("notes", ""),
                is_active=bool(body.get("is_active", True)),
            )
            return JsonResponse({"success": True, "id": vend.pk,
                                 "message": f"Vendor '{vend.vendor_name}' created."})
        except Exception as exc:
            logger.exception("api_config_vendors POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
@permission_required_code("procurement.view")
def api_config_vendor_detail(request, pk):
    """Get, Update, or Delete a Vendor."""
    from apps.procurement.models import Vendor
    from django.http import JsonResponse

    try:
        vend = Vendor.objects.get(pk=pk)
    except Vendor.DoesNotExist:
        return JsonResponse({"success": False, "message": "Vendor not found."}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": vend.pk,
            "vendor_name": vend.vendor_name,
            "country": vend.country,
            "city": vend.city,
            "address": vend.address,
            "contact_email": vend.contact_email,
            "contact_phone": vend.contact_phone,
            "preferred_vendor": vend.preferred_vendor,
            "reliability_score": float(vend.reliability_score),
            "average_lead_time_days": vend.average_lead_time_days,
            "payment_terms": vend.payment_terms,
            "bulk_discount_available": vend.bulk_discount_available,
            "rush_order_capable": vend.rush_order_capable,
            "on_time_delivery_pct": float(vend.on_time_delivery_pct),
            "notes": vend.notes,
            "is_active": vend.is_active,
        })

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            action = body.get("_action", "update")

            if action == "delete":
                name = vend.vendor_name
                vend.delete()
                return JsonResponse({"success": True, "message": f"Vendor '{name}' deleted."})

            if action == "toggle":
                vend.is_active = not vend.is_active
                vend.save(update_fields=["is_active"])
                return JsonResponse({"success": True, "is_active": vend.is_active,
                                     "message": f"Vendor '{vend.vendor_name}' {'activated' if vend.is_active else 'deactivated'}."})

            vend.vendor_name = body.get("vendor_name", vend.vendor_name).strip()
            vend.country = body.get("country", vend.country).strip()
            vend.city = body.get("city", vend.city).strip()
            vend.address = body.get("address", vend.address)
            vend.contact_email = body.get("contact_email", vend.contact_email)
            vend.contact_phone = body.get("contact_phone", vend.contact_phone)
            vend.preferred_vendor = bool(body.get("preferred_vendor", vend.preferred_vendor))
            vend.reliability_score = float(body.get("reliability_score", vend.reliability_score))
            vend.average_lead_time_days = int(body.get("average_lead_time_days", vend.average_lead_time_days))
            vend.payment_terms = body.get("payment_terms", vend.payment_terms)
            vend.bulk_discount_available = bool(body.get("bulk_discount_available", vend.bulk_discount_available))
            vend.rush_order_capable = bool(body.get("rush_order_capable", vend.rush_order_capable))
            vend.on_time_delivery_pct = float(body.get("on_time_delivery_pct", float(vend.on_time_delivery_pct)))
            vend.notes = body.get("notes", vend.notes)
            vend.is_active = bool(body.get("is_active", vend.is_active))
            vend.save()
            return JsonResponse({"success": True, "message": f"Vendor '{vend.vendor_name}' updated."})
        except Exception as exc:
            logger.exception("api_config_vendor_detail POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ---------------------------------------------------------------------------
# AJAX API -- Rooms (facility rooms)
# ---------------------------------------------------------------------------

@login_required
@permission_required_code("procurement.view")
def api_config_rooms(request):
    """List (GET) or Create (POST) Room entries."""
    from apps.procurement.models import Room
    from django.http import JsonResponse

    if request.method == "GET":
        q = request.GET.get("q", "").strip()
        qs = Room.objects.all().order_by("building_name", "floor_number", "room_code")
        if q:
            qs = qs.filter(
                Q(room_code__icontains=q) | Q(building_name__icontains=q) | Q(usage_type__icontains=q)
            )
        items = [
            {
                "id": rm.pk,
                "room_code": rm.room_code,
                "building_name": rm.building_name,
                "floor_number": rm.floor_number,
                "area_sqm": float(rm.area_sqm),
                "ceiling_height_m": float(rm.ceiling_height_m),
                "usage_type": rm.usage_type,
                "usage_type_display": rm.get_usage_type_display(),
                "design_temp_c": float(rm.design_temp_c),
                "temp_tolerance_c": float(rm.temp_tolerance_c),
                "design_cooling_load_kw": float(rm.design_cooling_load_kw),
                "design_humidity_pct": rm.design_humidity_pct,
                "noise_limit_db": rm.noise_limit_db,
                "contact_name": rm.contact_name,
                "contact_email": rm.contact_email,
                "is_active": rm.is_active,
            }
            for rm in qs
        ]
        return JsonResponse({"items": items, "total": len(items)})

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            rm = Room.objects.create(
                room_code=body.get("room_code", "").strip(),
                building_name=body.get("building_name", "").strip(),
                floor_number=int(body.get("floor_number", 0)),
                area_sqm=float(body.get("area_sqm", 50)),
                ceiling_height_m=float(body.get("ceiling_height_m", 3.0)),
                usage_type=body.get("usage_type", "OFFICE"),
                design_temp_c=float(body.get("design_temp_c", 22.0)),
                temp_tolerance_c=float(body.get("temp_tolerance_c", 1.0)),
                design_cooling_load_kw=float(body.get("design_cooling_load_kw", 5.0)),
                design_humidity_pct=int(body.get("design_humidity_pct")) if body.get("design_humidity_pct") else None,
                noise_limit_db=int(body.get("noise_limit_db")) if body.get("noise_limit_db") else None,
                location_description=body.get("location_description", ""),
                contact_name=body.get("contact_name", ""),
                contact_email=body.get("contact_email", ""),
                is_active=bool(body.get("is_active", True)),
            )
            return JsonResponse({"success": True, "id": rm.pk,
                                 "message": f"Room '{rm.room_code}' created."})
        except Exception as exc:
            logger.exception("api_config_rooms POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
@permission_required_code("procurement.view")
def api_config_room_detail(request, pk):
    """Get, Update, or Delete a Room."""
    from apps.procurement.models import Room
    from django.http import JsonResponse

    try:
        rm = Room.objects.get(pk=pk)
    except Room.DoesNotExist:
        return JsonResponse({"success": False, "message": "Room not found."}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": rm.pk,
            "room_code": rm.room_code,
            "building_name": rm.building_name,
            "floor_number": rm.floor_number,
            "area_sqm": float(rm.area_sqm),
            "ceiling_height_m": float(rm.ceiling_height_m),
            "usage_type": rm.usage_type,
            "design_temp_c": float(rm.design_temp_c),
            "temp_tolerance_c": float(rm.temp_tolerance_c),
            "design_cooling_load_kw": float(rm.design_cooling_load_kw),
            "design_humidity_pct": rm.design_humidity_pct or "",
            "noise_limit_db": rm.noise_limit_db or "",
            "location_description": rm.location_description,
            "contact_name": rm.contact_name,
            "contact_email": rm.contact_email,
            "is_active": rm.is_active,
        })

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            action = body.get("_action", "update")

            if action == "delete":
                code = rm.room_code
                rm.delete()
                return JsonResponse({"success": True, "message": f"Room '{code}' deleted."})

            if action == "toggle":
                rm.is_active = not rm.is_active
                rm.save(update_fields=["is_active"])
                return JsonResponse({"success": True, "is_active": rm.is_active,
                                     "message": f"Room '{rm.room_code}' {'activated' if rm.is_active else 'deactivated'}."})

            rm.room_code = body.get("room_code", rm.room_code).strip()
            rm.building_name = body.get("building_name", rm.building_name).strip()
            rm.floor_number = int(body.get("floor_number", rm.floor_number))
            rm.area_sqm = float(body.get("area_sqm", float(rm.area_sqm)))
            rm.ceiling_height_m = float(body.get("ceiling_height_m", float(rm.ceiling_height_m)))
            rm.usage_type = body.get("usage_type", rm.usage_type)
            rm.design_temp_c = float(body.get("design_temp_c", float(rm.design_temp_c)))
            rm.temp_tolerance_c = float(body.get("temp_tolerance_c", float(rm.temp_tolerance_c)))
            rm.design_cooling_load_kw = float(body.get("design_cooling_load_kw", float(rm.design_cooling_load_kw)))
            hum = body.get("design_humidity_pct")
            rm.design_humidity_pct = int(hum) if hum else None
            nse = body.get("noise_limit_db")
            rm.noise_limit_db = int(nse) if nse else None
            rm.location_description = body.get("location_description", rm.location_description)
            rm.contact_name = body.get("contact_name", rm.contact_name)
            rm.contact_email = body.get("contact_email", rm.contact_email)
            rm.is_active = bool(body.get("is_active", rm.is_active))
            rm.save()
            return JsonResponse({"success": True, "message": f"Room '{rm.room_code}' updated."})
        except Exception as exc:
            logger.exception("api_config_room_detail POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ---------------------------------------------------------------------------
# AJAX API -- HVAC Recommendation Rules
# ---------------------------------------------------------------------------

@login_required
@permission_required_code("procurement.view")
def api_config_hvacrules(request):
    """List (GET) or Create (POST) HVACRecommendationRule entries."""
    from apps.procurement.models import HVACRecommendationRule
    from django.http import JsonResponse

    if request.method == "GET":
        q = request.GET.get("q", "").strip()
        qs = HVACRecommendationRule.objects.all().order_by("priority", "rule_code")
        if q:
            qs = qs.filter(
                Q(rule_code__icontains=q) | Q(rule_name__icontains=q)
                | Q(recommended_system__icontains=q) | Q(rationale__icontains=q)
            )

        def _fmt(r):
            area_parts = []
            if r.area_sq_ft_min is not None:
                area_parts.append(f">= {r.area_sq_ft_min:,.0f}")
            if r.area_sq_ft_max is not None:
                area_parts.append(f"< {r.area_sq_ft_max:,.0f}")
            area_display = " & ".join(area_parts) if area_parts else "Any"
            temp_display = f">= {r.ambient_temp_min_c} C" if r.ambient_temp_min_c is not None else "Any"
            return {
                "id": r.pk,
                "rule_code": r.rule_code,
                "rule_name": r.rule_name,
                "country_filter": r.country_filter,
                "city_filter": r.city_filter,
                "store_type_filter": r.store_type_filter,
                "store_type_display": r.get_store_type_filter_display() if r.store_type_filter else "Any",
                "area_sq_ft_min": r.area_sq_ft_min,
                "area_sq_ft_max": r.area_sq_ft_max,
                "area_display": area_display,
                "ambient_temp_min_c": r.ambient_temp_min_c,
                "temp_display": temp_display,
                "budget_level_filter": r.budget_level_filter,
                "budget_display": r.get_budget_level_filter_display() if r.budget_level_filter else "Any",
                "energy_priority_filter": r.energy_priority_filter,
                "energy_priority_display": r.get_energy_priority_filter_display() if r.energy_priority_filter else "Any",
                "recommended_system": r.recommended_system,
                "recommended_system_display": r.get_recommended_system_display(),
                "alternate_system": r.alternate_system,
                "rationale": r.rationale,
                "priority": r.priority,
                "is_active": r.is_active,
                "notes": r.notes,
            }

        items = [_fmt(r) for r in qs]
        return JsonResponse({"items": items, "total": len(items)})

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            rule = HVACRecommendationRule.objects.create(
                rule_code=body.get("rule_code", "").strip(),
                rule_name=body.get("rule_name", "").strip(),
                country_filter=body.get("country_filter", ""),
                city_filter=body.get("city_filter", ""),
                store_type_filter=body.get("store_type_filter", ""),
                area_sq_ft_min=float(body["area_sq_ft_min"]) if body.get("area_sq_ft_min") not in (None, "") else None,
                area_sq_ft_max=float(body["area_sq_ft_max"]) if body.get("area_sq_ft_max") not in (None, "") else None,
                ambient_temp_min_c=float(body["ambient_temp_min_c"]) if body.get("ambient_temp_min_c") not in (None, "") else None,
                budget_level_filter=body.get("budget_level_filter", ""),
                energy_priority_filter=body.get("energy_priority_filter", ""),
                recommended_system=body.get("recommended_system", ""),
                alternate_system=body.get("alternate_system", ""),
                rationale=body.get("rationale", ""),
                priority=int(body.get("priority", 100)),
                is_active=bool(body.get("is_active", True)),
                notes=body.get("notes", ""),
            )
            return JsonResponse({
                "success": True, "id": rule.pk,
                "message": f"Rule '{rule.rule_code}' created successfully.",
            })
        except Exception as exc:
            logger.exception("api_config_hvacrules POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@login_required
@permission_required_code("procurement.view")
def api_config_hvacrule_detail(request, pk):
    """Get, Update, Toggle or Delete a single HVACRecommendationRule."""
    from apps.procurement.models import HVACRecommendationRule
    from django.http import JsonResponse

    try:
        rule = HVACRecommendationRule.objects.get(pk=pk)
    except HVACRecommendationRule.DoesNotExist:
        return JsonResponse({"success": False, "message": "Rule not found."}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": rule.pk,
            "rule_code": rule.rule_code,
            "rule_name": rule.rule_name,
            "country_filter": rule.country_filter,
            "city_filter": rule.city_filter,
            "store_type_filter": rule.store_type_filter,
            "area_sq_ft_min": rule.area_sq_ft_min,
            "area_sq_ft_max": rule.area_sq_ft_max,
            "ambient_temp_min_c": rule.ambient_temp_min_c,
            "budget_level_filter": rule.budget_level_filter,
            "energy_priority_filter": rule.energy_priority_filter,
            "recommended_system": rule.recommended_system,
            "alternate_system": rule.alternate_system,
            "rationale": rule.rationale,
            "priority": rule.priority,
            "is_active": rule.is_active,
            "notes": rule.notes,
        })

    if request.method == "POST":
        try:
            body = json.loads(request.body)
            action = body.get("_action", "update")

            if action == "delete":
                code = rule.rule_code
                rule.delete()
                return JsonResponse({"success": True, "message": f"Rule '{code}' deleted."})

            if action == "toggle":
                rule.is_active = not rule.is_active
                rule.save(update_fields=["is_active"])
                return JsonResponse({"success": True, "is_active": rule.is_active,
                                     "message": f"Rule '{rule.rule_code}' {'activated' if rule.is_active else 'deactivated'}."})

            rule.rule_code = body.get("rule_code", rule.rule_code).strip()
            rule.rule_name = body.get("rule_name", rule.rule_name).strip()
            rule.country_filter = body.get("country_filter", rule.country_filter)
            rule.city_filter = body.get("city_filter", rule.city_filter)
            rule.store_type_filter = body.get("store_type_filter", rule.store_type_filter)
            rule.area_sq_ft_min = float(body["area_sq_ft_min"]) if body.get("area_sq_ft_min") not in (None, "") else None
            rule.area_sq_ft_max = float(body["area_sq_ft_max"]) if body.get("area_sq_ft_max") not in (None, "") else None
            rule.ambient_temp_min_c = float(body["ambient_temp_min_c"]) if body.get("ambient_temp_min_c") not in (None, "") else None
            rule.budget_level_filter = body.get("budget_level_filter", rule.budget_level_filter)
            rule.energy_priority_filter = body.get("energy_priority_filter", rule.energy_priority_filter)
            rule.recommended_system = body.get("recommended_system", rule.recommended_system)
            rule.alternate_system = body.get("alternate_system", rule.alternate_system)
            rule.rationale = body.get("rationale", rule.rationale)
            rule.priority = int(body.get("priority", rule.priority))
            rule.is_active = bool(body.get("is_active", rule.is_active))
            rule.notes = body.get("notes", rule.notes)
            rule.save()
            return JsonResponse({"success": True, "message": f"Rule '{rule.rule_code}' updated."})
        except Exception as exc:
            logger.exception("api_config_hvacrule_detail POST failed: %s", exc)
            return JsonResponse({"success": False, "message": str(exc)}, status=400)

    return JsonResponse({"error": "Method not allowed"}, status=405)


# ---------------------------------------------------------------------------
# AJAX API -- HVAC Service Scope (read-only reference table)
# ---------------------------------------------------------------------------

@login_required
@permission_required_code("procurement.view")
def api_config_servicescopes(request):
    """List all HVACServiceScope rows (GET only -- managed via seed command)."""
    from apps.procurement.models import HVACServiceScope
    from django.http import JsonResponse

    if request.method != "GET":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    qs = HVACServiceScope.objects.all().order_by("sort_order", "system_type")
    items = [
        {
            "id": s.pk,
            "system_type": s.system_type,
            "system_type_display": s.display_name or s.system_type,
            "equipment_scope": s.equipment_scope,
            "installation_services": s.installation_services,
            "piping_ducting": s.piping_ducting,
            "electrical_works": s.electrical_works,
            "controls_accessories": s.controls_accessories,
            "testing_commissioning": s.testing_commissioning,
            "is_active": s.is_active,
        }
        for s in qs
    ]
    return JsonResponse({"items": items, "total": len(items)})


# ---------------------------------------------------------------------------
# External Suggestions API
# ---------------------------------------------------------------------------
@login_required
@permission_required_code("procurement.view")
def api_external_suggestions(request, pk):
    """Return AI-generated market intelligence with citations for this HVAC request.

    GET /procurement/<pk>/external-suggestions/
    Delegates to MarketIntelligenceService which handles LLM call, normalisation,
    and DB persistence.  Result is returned as JSON for the AJAX Refresh button.
    """
    from apps.procurement.services.market_intelligence_service import MarketIntelligenceService

    proc_request = get_object_or_404(ProcurementRequest, pk=pk)

    try:
        result = MarketIntelligenceService.generate_auto(
            proc_request,
            generated_by=request.user if request.user.is_authenticated else None,
        )
    except Exception as exc:
        logger.error(
            "api_external_suggestions LLM call failed for pk=%s: %s",
            pk, exc, exc_info=True,
        )
        _, system_code, system_name = MarketIntelligenceService.get_rec_context(proc_request)
        return JsonResponse({
            "system_code": system_code,
            "system_name": system_name,
            "rephrased_query": f"Market data for {system_name or 'HVAC system'} in {proc_request.geography_country or 'UAE'}",
            "ai_summary": f"Analysis failed: {exc}",
            "market_context": "",
            "suggestions": [],
            "error": str(exc),
        }, status=200)

    return JsonResponse(result)


# ---------------------------------------------------------------------------
# Market Intelligence page
# ---------------------------------------------------------------------------

@login_required
@permission_required_code("procurement.view")
def market_intelligence_page(request, pk):
    """AI Market Intelligence page for a procurement request.

    GET /procurement/<pk>/market-intelligence/
    Shows AI-generated product suggestions immediately from the DB (latest saved run).
    If none exist yet the page auto-triggers the LLM.  A Refresh button re-runs the LLM.
    """
    from apps.procurement.models import MarketIntelligenceSuggestion
    proc_request = get_object_or_404(ProcurementRequest, pk=pk)
    attributes = list(
        ProcurementRequestAttribute.objects
        .filter(request=proc_request)
        .values("attribute_label", "value_text", "value_number")
    )
    latest = (
        MarketIntelligenceSuggestion.objects
        .filter(request=proc_request)
        .order_by("-created_at")
        .first()
    )
    context = {
        "proc_request": proc_request,
        "attributes": attributes,
        "latest": latest,
        "page_title": f"Market Intelligence -- {proc_request.title}",
        "default_query": (
            f"Commercial HVAC system procurement for "
            f"{proc_request.geography_country or 'UAE'} retail facility: "
            f"{proc_request.title}. "
            f"Location: {proc_request.geography_city or ''}, "
            f"{proc_request.geography_country or 'UAE'}. "
            f"Find real 2025/2026 product options, AED pricing, and GCC distributor availability."
        ),
    }
    return render(request, "procurement/market_intelligence.html", context)


@login_required
@permission_required_code("procurement.view")
def api_perplexity_research(request, pk):
    """AJAX: run a live Perplexity sonar-pro web search for market intelligence."""
    proc_request = get_object_or_404(ProcurementRequest, pk=pk)
    custom_query = request.GET.get("query", "").strip()

    # Build context blocks from DB
    attributes = list(
        ProcurementRequestAttribute.objects
        .filter(request=proc_request)
        .values("attribute_code", "attribute_label", "value_text", "value_number")
    )
    attr_lines = []
    for a in attributes:
        val = a["value_text"] or (str(a["value_number"]) if a["value_number"] else "")
        if val:
            attr_lines.append(f"  - {a['attribute_label']}: {val}")
    attrs_block = "\n".join(attr_lines) if attr_lines else "  (no attributes recorded)"

    recommendation = (
        RecommendationResult.objects
        .filter(run__request=proc_request)
        .order_by("-created_at")
        .first()
    )
    system_code = ""
    rec_block = "(no internal recommendation yet)"
    if recommendation:
        payload = recommendation.output_payload_json or {}
        system_code = payload.get("system_type_code", "")
        conf = int((recommendation.confidence_score or 0) * 100)
        rec_block = f"Recommended: {system_code} (confidence {conf}%)"

    # Load approved sources restricted to this product's HVAC system type
    from apps.procurement.models import ExternalSourceRegistry as _ESR
    _SOURCE_CLASS_ORDER = {"OEM_OFFICIAL": 0, "AUTHORIZED_DISTRIBUTOR": 1, "OEM_REGIONAL": 2}
    _approved_sources = []
    if system_code:
        _approved_sources = list(
            _ESR.objects.filter(
                hvac_system_type=system_code,
                is_active=True,
                allowed_for_discovery=True,
            ).values("source_name", "domain", "source_url", "source_class")
        )
    if not _approved_sources:
        # Fallback: all active discovery-enabled sources (no open-web leak)
        _approved_sources = list(
            _ESR.objects.filter(
                is_active=True,
                allowed_for_discovery=True,
            ).values("source_name", "domain", "source_url", "source_class")
        )
    _approved_sources.sort(key=lambda s: _SOURCE_CLASS_ORDER.get(s["source_class"], 99))
    _domain_list = [s["domain"] for s in _approved_sources]
    _domain_to_url = {
        s["domain"]: (s["source_url"] or f"https://{s['domain']}")
        for s in _approved_sources
    }
    _sources_lines = [
        f"  - {s['source_name']} | {s['source_url'] or 'https://' + s['domain']} | domain: {s['domain']}"
        for s in _approved_sources
    ]
    _sources_block = "\n".join(_sources_lines)

    api_key = getattr(settings, "PERPLEXITY_API_KEY", "")
    model = getattr(settings, "PERPLEXITY_MODEL", "sonar-pro")

    if not api_key:
        return JsonResponse(
            {"error": "Perplexity API key is not configured in settings."},
            status=500,
        )

    # ---- Build prompts ----
    SYSTEM_PROMPT = (
        "You are a senior HVAC market intelligence analyst specializing in commercial "
        "and retail HVAC systems for the GCC/Middle East region. "
        "You have live web search access. Research real current manufacturer product lines, "
        "pricing in AED, regional distributor availability, and compliance with local standards "
        "(ESMA, ASHRAE). Focus on brands active in the region: "
        "Daikin, Carrier, Trane, York, Mitsubishi Electric, LG, Samsung, Gree, Midea, Voltas. "
        "Respond ONLY with a single valid JSON object and nothing else."
    )

    research_focus = custom_query if custom_query else (
        f"Commercial HVAC system procurement for a "
        f"{proc_request.geography_country or 'UAE'} retail/commercial facility. "
        f"Facility: {proc_request.title}. "
        f"Location: {proc_request.geography_city or ''}, {proc_request.geography_country or 'UAE'}. "
        f"{proc_request.description or ''}"
    )

    USER_PROMPT = f"""Research this HVAC procurement need using live web data and return exact JSON.

=== PROCUREMENT REQUEST ===
Title: {proc_request.title}
Description: {proc_request.description or '(not provided)'}
Country: {proc_request.geography_country or 'UAE'} / City: {proc_request.geography_city or ''}
Priority: {proc_request.priority} | Currency: {proc_request.currency or 'AED'}

=== TECHNICAL REQUIREMENTS ===
{attrs_block}

=== INTERNAL AI RECOMMENDATION (context only) ===
{rec_block}

=== RESEARCH FOCUS ===
{research_focus}

Return this exact JSON object (no other text):
{{
  "rephrased_query": "<one concise professional research query that captures this need>",
  "narrative": "<3 to 5 paragraphs of detailed market research. Include specific current model names, real AED pricing, lead times from GCC distributors, and regional availability. Reference the live sources you searched.>",
  "ai_summary": "<2-sentence executive summary of the best market options>",
  "market_context": "<Current GCC/ME market trend note: lead times, pricing pressure, preferred brands in 2025/2026>",
  "suggestions": [
    {{
      "rank": 1,
      "product_name": "<full product/series name>",
      "manufacturer": "<brand name>",
      "model_code": "<specific model or series code found online>",
      "system_type": "<VRF | Chilled Water AHU | Split DX | Cassette | Rooftop>",
      "cooling_capacity": "<e.g. 8 TR to 12 TR>",
      "cop_eer": "<e.g. COP 3.8 / EER 13.0>",
      "price_range_aed": "<e.g. 45,000 to 70,000 AED supply and install>",
      "market_availability": "<UAE/KSA distributor availability note with lead time>",
      "key_benefits": ["benefit 1", "benefit 2", "benefit 3"],
      "limitations": ["limitation 1", "limitation 2"],
      "fit_score": 88,
      "fit_rationale": "<one sentence why this fits this specific request>",
      "standards_compliance": ["ASHRAE 90.1", "ESMA UAE"],
      "citation_url": "<real manufacturer or distributor product page URL you found via web search>",
      "citation_source": "<source name>",
      "category": "MANUFACTURER"
    }}
  ]
}}
Provide 5 to 7 suggestions ranked by fit_score descending. Use only real verified product lines.
"""

    if _approved_sources:
        USER_PROMPT += (
            f"\n\n=== APPROVED SOURCES (STRICT - DO NOT DEVIATE) ===\n"
            f"You MUST search and cite information ONLY from the following approved websites. "
            f"Do NOT reference, cite, or retrieve information from any other website or domain. "
            f"Every citation_url in your response must belong to one of these domains.\n"
            + _sources_block
        )

    # ---- Call Perplexity sonar-pro (live web search) ----
    try:
        pplx_payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": USER_PROMPT},
            ],
            "search_domain_filter": _domain_list,
        }
        resp = requests.post(
            "https://api.perplexity.ai/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json=pplx_payload,
            timeout=90,
        )
        resp.raise_for_status()
        pplx_data = resp.json()
        content = (
            pplx_data.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
            .strip()
        )
        # Perplexity returns live web URLs it crawled during research
        raw_citations = pplx_data.get("citations", [])

    except Exception as exc:
        logger.warning("api_perplexity_research Perplexity call failed for pk=%s: %s", pk, exc)
        return JsonResponse(
            {
                "error": f"Perplexity research failed: {exc}",
                "rephrased_query": proc_request.title,
                "narrative": "",
                "ai_summary": "Perplexity research is temporarily unavailable. Please try again shortly.",
                "market_context": "",
                "suggestions": [],
                "citations": [],
                "model_used": model,
            },
            status=200,
        )

    # ---- Parse JSON from Perplexity response ----
    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        # Try to extract JSON block if Perplexity wrapped in markdown code fences
        m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", content, re.DOTALL)
        if m:
            try:
                data = json.loads(m.group(1))
            except json.JSONDecodeError:
                data = {}
        else:
            data = {"narrative": content}

    # ---- Normalise suggestion fit scores ----
    suggestions = data.get("suggestions", [])
    for s in suggestions:
        try:
            s["fit_score"] = max(0, min(100, int(s.get("fit_score", 0))))
        except (TypeError, ValueError):
            s["fit_score"] = 0

    # ---- Enforce approved-domain citations ----
    if _approved_sources and _domain_list:
        def _ci_dom(url):
            try:
                return urlparse(url).netloc.lower().lstrip("www.")
            except Exception:
                return ""
        for s in suggestions:
            curl = s.get("citation_url") or ""
            if curl:
                cd = _ci_dom(curl)
                matched = any(
                    cd == ad.lower().lstrip("www.") or
                    cd.endswith("." + ad.lower().lstrip("www."))
                    for ad in _domain_list
                )
                if not matched:
                    csrc = (s.get("citation_source") or "").lower()
                    replacement = None
                    for src in _approved_sources:
                        if src["source_name"].lower() in csrc or csrc in src["source_name"].lower():
                            replacement = src["source_url"] or f"https://{src['domain']}"
                            break
                    s["citation_url"] = replacement or _domain_to_url.get(_domain_list[0], f"https://{_domain_list[0]}")
            else:
                s["citation_url"] = _domain_to_url.get(_domain_list[0], f"https://{_domain_list[0]}")

    # ---- Build citation cards -- only from approved domains ----
    _approved_domain_bare = {ad.lower().lstrip("www.") for ad in _domain_list}

    def _is_approved_url(url):
        try:
            host = urlparse(url).netloc.lower().lstrip("www.")
            return any(
                host == ad or host.endswith("." + ad)
                for ad in _approved_domain_bare
            )
        except Exception:
            return False

    citation_cards = []
    seen_domains: set = set()
    for url in raw_citations:
        if not _is_approved_url(url):
            continue
        try:
            domain = urlparse(url).netloc.replace("www.", "")
        except Exception:
            domain = url
        if domain not in seen_domains:
            seen_domains.add(domain)
            citation_cards.append({"url": url, "domain": domain})

    # Also add per-suggestion citation URLs (already restricted above) if not yet in list
    for s in suggestions:
        curl = s.get("citation_url", "")
        if curl and _is_approved_url(curl):
            try:
                cdomain = urlparse(curl).netloc.replace("www.", "")
            except Exception:
                cdomain = curl
            if cdomain not in seen_domains:
                seen_domains.add(cdomain)
                citation_cards.append({"url": curl, "domain": cdomain})

    return JsonResponse({
        "rephrased_query": data.get("rephrased_query", proc_request.title),
        "narrative": data.get("narrative", content),
        "ai_summary": data.get("ai_summary", ""),
        "market_context": data.get("market_context", ""),
        "suggestions": suggestions,
        "citations": citation_cards,
        "model_used": model,
    })


# ---------------------------------------------------------------------------
# Stores Management (full CRUD page)
# ---------------------------------------------------------------------------

@login_required
@permission_required_code("procurement.view")
def stores_management(request):
    """Full-page store management: list, search, create, edit, toggle status."""
    query = (request.GET.get("q") or "").strip()
    country_filter = (request.GET.get("country") or "").strip()
    store_type_filter = (request.GET.get("store_type") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()

    qs = HVACStoreProfile.objects.all().order_by("store_id")
    if query:
        qs = qs.filter(
            Q(store_id__icontains=query)
            | Q(brand__icontains=query)
            | Q(city__icontains=query)
            | Q(country__icontains=query)
        )
    if country_filter:
        qs = qs.filter(country__iexact=country_filter)
    if store_type_filter:
        qs = qs.filter(store_type__iexact=store_type_filter)
    if status_filter == "active":
        qs = qs.filter(is_active=True)
    elif status_filter == "inactive":
        qs = qs.filter(is_active=False)

    total_stores = HVACStoreProfile.objects.count()
    active_stores = HVACStoreProfile.objects.filter(is_active=True).count()
    inactive_stores = total_stores - active_stores

    countries = (
        HVACStoreProfile.objects
        .exclude(country="")
        .values_list("country", flat=True)
        .distinct()
        .order_by("country")
    )
    store_types = (
        HVACStoreProfile.objects
        .exclude(store_type="")
        .values_list("store_type", flat=True)
        .distinct()
        .order_by("store_type")
    )

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "procurement/stores_management.html", {
        "page_obj": page_obj,
        "query": query,
        "country_filter": country_filter,
        "store_type_filter": store_type_filter,
        "status_filter": status_filter,
        "total_stores": total_stores,
        "active_stores": active_stores,
        "inactive_stores": inactive_stores,
        "countries": list(countries),
        "store_types": list(store_types),
    })


@login_required
@permission_required_code("procurement.create")
def api_store_management_create(request):
    """AJAX POST -- create a new store from the Stores Management page."""
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    store_id = (request.POST.get("store_id") or "").strip()
    if not store_id:
        return JsonResponse({"error": "Store ID is required."}, status=400)

    if HVACStoreProfile.objects.filter(store_id=store_id).exists():
        return JsonResponse({"error": f"Store ID '{store_id}' already exists."}, status=409)

    try:
        defaults = _build_hvac_store_profile_defaults(request.POST)
        profile = HVACStoreProfile.objects.create(
            store_id=store_id,
            created_by=request.user,
            **defaults,
        )
        return JsonResponse({"ok": True, "store_id": profile.store_id, "pk": profile.pk}, status=201)
    except Exception as exc:
        logger.exception("api_store_management_create failed: %s", exc)
        return JsonResponse({"error": str(exc)}, status=500)


@login_required
@permission_required_code("procurement.create")
def api_store_management_detail(request, pk):
    """AJAX: GET detail, POST update, POST with _action=DELETE to remove."""
    profile = get_object_or_404(HVACStoreProfile, pk=pk)

    if request.method == "GET":
        return JsonResponse({
            "pk": profile.pk,
            "store_id": profile.store_id,
            "brand": profile.brand,
            "country": profile.country,
            "city": profile.city,
            "store_type": profile.store_type,
            "store_format": profile.store_format,
            "area_sqft": profile.area_sqft,
            "ceiling_height_ft": profile.ceiling_height_ft,
            "operating_hours": profile.operating_hours,
            "footfall_category": profile.footfall_category,
            "ambient_temp_max": profile.ambient_temp_max,
            "humidity_level": profile.humidity_level,
            "dust_exposure": profile.dust_exposure,
            "heat_load_category": profile.heat_load_category,
            "fresh_air_requirement": profile.fresh_air_requirement,
            "landlord_constraints": profile.landlord_constraints,
            "existing_hvac_type": profile.existing_hvac_type,
            "budget_level": profile.budget_level,
            "energy_efficiency_priority": profile.energy_efficiency_priority,
            "is_active": profile.is_active,
        })

    if request.method == "POST":
        action = (request.POST.get("_action") or "").strip().upper()
        if action == "DELETE":
            store_id = profile.store_id
            profile.delete()
            return JsonResponse({"ok": True, "deleted_pk": pk, "store_id": store_id})

        # UPDATE path
        defaults = _build_hvac_store_profile_defaults(request.POST)
        for field, value in defaults.items():
            setattr(profile, field, value)
        is_active_val = request.POST.get("is_active")
        if is_active_val is not None:
            profile.is_active = is_active_val in ("true", "1", "True", "on")
        try:
            profile.save()
            return JsonResponse({"ok": True, "pk": profile.pk, "store_id": profile.store_id})
        except Exception as exc:
            logger.exception("api_store_management_detail update failed: %s", exc)
            return JsonResponse({"error": str(exc)}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)
